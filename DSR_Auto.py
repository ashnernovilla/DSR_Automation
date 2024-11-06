# %% [markdown]
# # Life Lines Automate

# %%
import os
from time import sleep
import pyautogui
import subprocess
import ctypes
import glob
import os, shutil
from datetime import datetime
# import pygetwindow as gw
from pywinauto.application import Application
from pywinauto.keyboard import send_keys
from pathlib import Path

# passvalue = "AiaBpi@789"
passvalue = "Open@1234"

## 1. Please create a YTDN Dump in Documents named YTDN_Dump
## 2. Please add a shortcut of LifeLines in Downloads
## 3. Make Sure yuor Chrome Drivers are updated
## 4. 

path = r'\\10.117.187.29\bplac\BPLPDCBG'
ytdn = path + '\\YTDN-01.TXT'

shared_folder = Path(path)


username = "BPLP460A"
password = passvalue

# %% [markdown]
# ## Running the YTDS

# %%
os.chdir(r'C:\Users\i024605\OneDrive - AIA Group Ltd\Documents\PythonCodes\Python_Pyfiles\DSR_Automate')
print(os.getcwd())

import YTDS as ytds
ytds.main(password)


# %%
os.chdir(r'\\Palp3r7cfis08\bplac mis library\Git_Codes\DSR_Codes')
print(os.getcwd())

# %%
cmd = f'net use {shared_folder} /user:{username} {password}'
# os.system(cmd)

os.listdir(path)

## First Enter the Path of Lifelines as well its shrtcut name
os.chdir(r'C:\Users\I024605\Documents')
# Specify the path to the shortcut or script that launches the application
shortcut_path ="LifeLines.lnk" # Replace with the actual path


name_date_now = datetime.now().strftime("%m%Y")

# local_path = r'C:\Users\I024605\Documents\YTDN_Dump'
local_path = os.path.join(os.path.expanduser("~"), "Documents\YTDN_Dump")

source = ytdn
destination = local_path+f"\\YTDN-01-{name_date_now}.TXT"

print(source, destination)

sleep(5)

try:
    dest = shutil.move(source, destination) 
except:
    print("No File")
    pass


sleep(5)

## 2nd Import the subprocess and run the lifelines
lifeline_process = subprocess.Popen(shortcut_path, shell=True)  # returns the exit code in unix

# Wait for 10 Seconds because the opening is slow
sleep(10)


## Import ctypes to make the life lins on top of other applications
user32 = ctypes.windll.user32
lifeline_handle = user32.FindWindowW(None, "A - 5250 Display")

ctypes.windll.user32.ShowWindow(lifeline_handle, 3)

while lifeline_handle == 0:
    user32 = ctypes.windll.user32
    lifeline_handle = user32.FindWindowW(None, "A - 5250 Display")
    if lifeline_handle != 0:
        ctypes.windll.user32.ShowWindow(lifeline_handle, 3)
        break
    else:
        print("Ongoing Loading")
        sleep(1)

sleep(8)

try:
    app=Application
    app = Application(backend="win32").connect(title_re='Signon to IBM i', timeout=5)
    form = app.window(title_re='Signon to IBM i')
    form.send_keystrokes (password)
    form.send_keystrokes ('{TAB}')
    form.send_keystrokes ('{ENTER}')
    sleep(10)
except:
    print("pass")
    pass

app=Application
app = Application(backend="win32").connect(title_re='A - 5250 Display', timeout=5)
form = app.window(title_re='A - 5250 Display')

## Loggin in the System
sleep(8)
form.send_keystrokes ('BPLP460A')
form.send_keystrokes ('{TAB}')
form.send_keystrokes (password)
form.send_keystrokes ('{ENTER}')
form.send_keystrokes ('{ENTER}')
sleep(1)

## Manila Palas400
form.send_keystrokes ('1')
form.send_keystrokes ('{ENTER}')
sleep(1)

## LifeLines Session Start
form.send_keystrokes ('1')
form.send_keystrokes ('{ENTER}')
sleep(1)

## Main Selection Menu LLPM00
form.send_keystrokes ('13')
form.send_keystrokes ('{ENTER}')
form.send_keystrokes ('{ENTER}')
sleep(1)

## Main Selection Menu LLPM00 Download Files
form.send_keystrokes ('19')
form.send_keystrokes ('{ENTER}')
form.send_keystrokes ('{ENTER}')
sleep(1)

## File Download Request Menu
form.send_keystrokes ('1')
form.send_keystrokes ('{ENTER}')
form.send_keystrokes ('{ENTER}')
sleep(1)

## Bplac Sales Operations
form.send_keystrokes ('7')
form.send_keystrokes ('{ENTER}')
form.send_keystrokes ('{ENTER}')
sleep(1)

while True:
   
    # list_of_files = glob.glob(path+'\\YTDN-01.TXT') # * means all if need specific format then *.
    # path = r'C:\Users\I024605\Documents\YTDN_Dump'
    list_of_files = glob.glob(path+'\\YTDN-01.TXT')
    
    try:    
        if os.path.basename(list_of_files[0]) == 'YTDN-01.TXT':
            print('File Exist')
            
            os.system('taskkill /F /IM acslaunch_win-64.exe')
            
            break
            
        else:
            print("File is not Existing")
            sleep(120)
            continue
    except:
        print("File is not Existing")
        sleep(120)
        continue
    
    sleep(120)

# %%
import os
import win32com.client as win32
import glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime as dt
from pathlib import Path

try:

    path_file_upload = os.getcwd()

    path = r'\\10.117.187.29\bplac\BPLPDCBG'
    ytdn = path + '\\YTDN-01.TXT'

    shared_folder = Path(path)


    username = "BPLP460A"
    password = password

    cmd = f'net use {shared_folder} /user:{username} {password}'
    os.system(cmd)

    os.listdir(path)

    Date = dt.now().strftime('%B %d, %Y')

    olApp = win32.Dispatch('Outlook.Application')
    olNS = olApp.GetNamespace("MAPI")

    mailItem = olApp.CreateItem(0)
    mailItem.Subject = f'YTDN - {Date}'
    mailItem.BodyFormat = 1
    mailItem.HTMLBody = f'''
    <!DOCTYPE html>
        <html>
        <body>
        <div><p>Dear All,</p></div>
        <div><p>Please see YTDN File Attached.</p></div>
        <div>
            <p> YTDN Run Only - In a Few Minutes Daily Submission Report Will be Generated Current {Date}.
            </p>
        </div>

    <div><p>Best Regards,</p></div>
    </body>
    </html>
    ''' 
    #mailItem.To = 'ashnergerald.novilla@aia.com'
    mailItem.To = 'joshuaaudie-ja.depositario@aia.com'
    mailItem.cc = 'jonathan-j.tomas@aia.com'
    mailItem.Attachments.Add(ytdn)
    
    mailItem.SentOnBehalfOfName = "philippines.bplac.bas.incentive-prod@aia.com"

    mailItem.Save()
    mailItem.Send()
except:
    pass


# manual run
# %% [markdown]
# # MTDREF

# %% [markdown]
# ## Import All Python Package Needed

# %%
# import mitosheet 
# from mitosheet.public.v3 import *;

import pandas as pd
import numpy as np
from sqlalchemy import create_engine
import pyodbc 
from datetime import datetime, date, timedelta
import warnings
from dateutil.relativedelta import relativedelta

pd.set_option('display.max_columns', None)
warnings.filterwarnings('ignore')

# %% [markdown]
# ## Path of the File Needed

# %%
#path = r'D:\Documents'
#path = r'\\pas4ps01\BPLAC\BPLPDCBG'
#path = r'C:\Users\I024605\Documents\YTDN_Dump'
path = r'\\10.117.187.29\bplac\BPLPDCBG'
# path = r'C:\Users\I024605\Documents\YTDN_Dump'
ytdn = path + '\\YTDN-01.TXT'
#ytdn = path + '\\YTDN-01-102024.TXT'

# %% [markdown]
# ## Function for Trimming

# %%
def strip_element(x):
    if isinstance(x, str):
        return x.strip()
    return x

# %% [markdown]
# ## SQL Function For Selecting in Database

# %%
def sql_connection(sql_query, dbase):
    # Replace with your actual SQL Server connection details
    server = 'PPBWDLC0SG7A1'
    database = dbase
    username = 'admin'
    password = 'Openlab@123'

    # Define the connection string
    connection_string = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
    )


    try:
        # Establish the connection
        conn = pyodbc.connect(connection_string)

        # Define the SELECT query
        query = sql_query

        # Execute the query and fetch data into a DataFrame
        df = pd.read_sql(query, conn)

        # Execute Trims 
        df = df.applymap(strip_element)

        # Print the data
        print(df)


    except pyodbc.Error as e:
        print("Error connecting to SQL Server:", e)
    finally:
        # Close the connection
        conn.close()

    return df

# %% [markdown]
# ## Loading the YTDN Data 

# %%
dtype = {
    'COMPANY' : str,
    'SUBMISSION DATE' : str,
    'SUBMISSION TIME' : str,
    'POLICY NO.' : str,
    'APPROVAL DATE' : str,
    'ISSUE DATE' : str,
    'STATUS CODE' : str,
    'BILL MODE' : str,
    'BRANCH CODE' : str,
    'BSE CODE' : str,
    'BAM CODE' : str,
    'BDM CODE' : str,
    'TSH CODE' : str,
    'APPLICATION DATE' : str,
    'REFERRER CODE' : str
}

ytdn_raw = pd.read_csv(ytdn,encoding='ISO-8859-1', dtype=dtype)
# mitosheet.sheet(ytdn_raw, analysis_to_replay="id-fdhjfvirwj")
ytdn_raw

#os.system(f'net use {shared_folder} /delete')

# %% [markdown]
# ## Setting the Dollar Value

# %%
# Change this when dollar value changed
Dollar_Value = 56.79

# %% [markdown]
# ## Fixing the Dates, Trim Spaces ,Renaming Company, Policy, Name, Suspense Amount, Droping Not Needed Columns

# %%
ytdn_1= ytdn_raw.copy()

ytdn_1 = ytdn_1.applymap(strip_element)


# Set formula of ISSUE DATE
ytdn_1['ISSUE DATE'] = '20'+ytdn_1['ISSUE DATE']

# Changed APPROVAL DATE to dtype datetime
ytdn_1['APPROVAL DATE'] = pd.to_datetime(ytdn_1['APPROVAL DATE'], format='%Y%m%d', errors='coerce')

# Changed ISSUE DATE to dtype datetime
ytdn_1['ISSUE DATE'] = pd.to_datetime(ytdn_1['ISSUE DATE'], format='%Y%m%d', errors='coerce')

# Changed APPLICATION DATE to dtype datetime
ytdn_1['APPLICATION DATE'] = pd.to_datetime(ytdn_1['APPLICATION DATE'], format='%Y%m%d', errors='coerce')


# Changed SUBMISSION DATE to dtype datetime
ytdn_1['SUBMISSION DATE'] = pd.to_datetime(ytdn_1['SUBMISSION DATE'], infer_datetime_format=True, errors='coerce')

# Add Column 'EXTRACTED DATE' date TODAY
ytdn_1['EXTRACTED DATE'] = datetime.today()

# Renamed columns CA_CO, CA_POLICY NO., CA_NAME, SUSAMT1
ytdn_1.rename(columns={
    'COMPANY': 'CA_CO',
    'POLICY NO.': 'CA_POLICY NO.',
    'NAME': 'CA_NAME',
    'SUSPENSE AMOUNT': 'SUSAMT1'
}, inplace=True)

# Deleted columns PAYMENT METHOD, ENROLLMENT STATUS, BANK ACCOUNT
ytdn_1.drop(['PAYMENT METHOD', 'ENROLLMENT STATUS', 'BANK ACCOUNT'], axis=1, inplace=True)

# Added column PHASE
ytdn_1.insert(17, 'PHASE', "")

# Added column SUBPHASE
ytdn_1.insert(18, 'SUBPHASE', 0)

# Deleted columns BRANCH NAME, BAM CODE, BAM NAME, BDM CODE, BDM NAME, TSH CODE, TSH NAME, ISSUING OFFICE, ORIGINATION - DESK, ORIGINATION - DEPARTMENT, SERVICING OFFICE, ISSUE AGE, APPLICATION DATE, REFERRER CATEGORY, MODAL PREMIUM
ytdn_1.drop(['BRANCH NAME', 'BAM CODE', 'BAM NAME', 'BDM CODE', 'BDM NAME', 'TSH CODE', 'TSH NAME', 'ISSUING OFFICE', 'ORIGINATION - DESK', 'ORIGINATION - DEPARTMENT', 'SERVICING OFFICE', 'ISSUE AGE', 'APPLICATION DATE', 'REFERRER CATEGORY', 'MODAL PREMIUM'], axis=1, inplace=True)

ytdn_1

# %% [markdown]
# ## Load the BSE Database

# %%
query = '''
WITH CorpsoloftheMonth AS
(
SELECT * FROM db_MasterFile_2024.[dbo].tbl_CorpsolMasterFile_2024 AS COR24
LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(COR24.WWDate,3) 
),
CORPS24 AS
(
SELECT COR24.* FROM CorpsoloftheMonth AS COR24 WHERE COR24.[INDEX] = (SELECT MAX(cor.[INDEX]) FROM CorpsoloftheMonth cor)
),
BSEoftheMonth AS 
( 
Select BSE24.*, mlu.[Index]  FROM db_MasterFile_2024.[dbo].tbl_BSEMasterFile_2024 AS BSE24 
LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(BSE24.WWDate,3) 
),
BSES24 AS
(
SELECT BSE24.* FROM BSEoftheMonth AS BSE24 WHERE BSE24.[INDEX] = (SELECT MAX(bse.[INDEX]) FROM BSEoftheMonth bse)
),
BSECORP AS
(
SELECT 
       BSES24.[POSITION]
      ,BSES24.[EE CODE]
      ,BSES24.[BSE FULL NAME]
      ,BSES24.[BSE CODE]
      ,BSES24.[SURNAME]
      ,BSES24.[FIRST NAME]
      ,BSES24.[MIDDLE NAME]
      ,BSES24.[SEGMENT]
      ,BSES24.[HYBRID]
      ,BSES24.[JUNIOR/ SENIOR]
      ,BSES24.[OLD/NEW]
      ,BSES24.[TENURE  (CURRENT POSITION)]
      ,BSES24.[DEPLOYMENT DATE (MM/DD/YYYY)]
      ,BSES24.[DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)]
      ,BSES24.[MOBILE NUMBERS]
      ,BSES24.[EMAILADD]
      ,CASE
		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
		ELSE BSES24.[BPI-AIA AREA]
	   END AS [BPI-AIA AREA]
      ,CASE
		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
		ELSE BSES24.[BPI-AIA DIVISION]
	   END AS [BPI-AIA DIVISION]
      ,CASE
		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
		ELSE BSES24.[TERRITORY]
	   END AS [TERRITORY]
      ,BSES24.[BAM]
      ,BSES24.[BAM CODE]
      ,BSES24.[BDM]
      ,BSES24.[BDM CODE]
      ,BSES24.[TSH]
      ,BSES24.[TSH CODE]
      ,BSES24.[TL NAME]
      ,BSES24.[TL CODE]
      ,BSES24.[WWDate]
      ,BSES24.[BATCH]
FROM BSES24

UNION ALL

SELECT 
	 CORPS24.[POSITION]	
	,CORPS24.[EE CODE]
	,CORPS24.[FULL NAME]
	,CORPS24.[POSITION CODE]
	,CORPS24.[SURNAME]	
	,CORPS24.[FIRST NAME]	
	,CORPS24.[MIDDLE NAME]
	,CORPS24.[SEGMENT]
	,'Hybrid' = NULL
	,'JUNIOR/ SENIOR' = NULL
	,CORPS24.[OLD/NEW]
	,CORPS24.[TENURE  (CURRENT POSITION)]
	,CORPS24.[DEPLOYMENT DATE]
	,'DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)' = NULL
	,CORPS24.[MOBILE NUMBERS]
	,CORPS24.[EMAILADD]
	,CASE
		WHEN [POSITION] IN ('MBBSA', 'CBSS') THEN 'CORPSOL'
		ELSE CORPS24.[BPLAC AREA]
	END AS [BPLAC AREA]
	,CASE
		WHEN [POSITION] IN ('MBBSA', 'CBSS') THEN 'CORPSOL'
		ELSE CORPS24.[BPLAC DIV]
	END AS [BPLAC DIV]
	,'TERRITORY'= 'CORPSOL'
	,CORPS24.[CSM] AS [BAM]
	,CORPS24.[CSM CODE] AS [BAM CODE]
	,CORPS24.[HEAD] AS [BDM]
	,CORPS24.[CHANNEL HEAD CODE] AS [BDM CODE]
	,'TSH' = NULL
	,'TSH CODE' = NULL
	,'TL NAME' = NULL
	,'TL CODE' = NULL
	,CORPS24.[WWDate]
	,'BATCH' = NULL
FROM CORPS24
WHERE [POSITION] IN ('MBBSA', 'CBSS')
)
SELECT * FROM BSECORP
'''
db_bse_masterfile = sql_connection(query, 'db_MasterFile_2024')
db_bse_masterfile

# %% [markdown]
# ## Load the MasterFile Database

# %%
query = '''
        WITH MfileoftheMonth AS 
    ( 
		Select 
		[RELA CODE]
      ,[VALIDATION CODE]
      ,[BANK  CODE]
      ,[BPI-AIA  BRANCH CODE]
      ,[BRANCH CODE  FOR BPI-AIA]
      ,[CLUSTER CODE]
      ,[BANK]
      ,[BRANCH NAME]
      ,[DIVISION]
      ,[AREA]
      ,[BRANCH TYPE]
      ,[BUSINESS  MANAGER CODE]
      ,[BUSINESS MANAGER]
      ,[ASST BUSINESS MANAGER]
      ,[BUSINESS DIRECTOR]
      ,[BPI DIVISION HEAD]
      ,[RM REF CODE]
      ,[RM]
      ,[SA]
      ,[SO]
      ,[BRANCH_ADD]
      ,[BRANCH_TEL]
      ,[TEL 2]
      ,[TEL 3]
      ,[BRANCH_FAX]
      ,[ZIP]
      ,[TERRITORY  CODE]
      ,[TERRITORY]
      ,[BPI-AIA  DIVISION CODE]
      ,[BPI-AIA DIVISION]
      ,[BPI-AIA  AREA CODE]
      ,[BPI-AIA AREA]
      ,[PERSONAL BSE 1]
      ,[PER1 CODE]
      ,[PERSONAL BSE 2]
      ,[PER2 CODE]
      ,[PERSONAL BSE 3]
      ,[PER3 CODE]
      ,[PERSONAL BSE 4]
      ,[PER4 CODE]
      ,[PERSONAL BSE 5]
      ,[PER5 CODE]
      ,[PREFERRED BSE 1]
      ,[PRF1 CODE]
      ,[PREFERRED BSE 2]
      ,[PRF2 CODE]
      ,[PREFERRED BSE 3]
      ,[PRF3 CODE]
      ,[BAM]
      ,[BAM CODE]
      ,[BDM]
      ,[BDM CODE]
      ,[TSH]
      ,[TSH CODE]
      ,[BRANCH OLD NAME]
      ,[STATUS OF BRANCHES BASED ON BSC HIERARCHY]
      ,[WWDate]
		, mlu.[Index]  FROM db_MasterFile_2024.[dbo].tbl_MasterFile_2024 AS Mfile24 
		LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(Mfile24.WWDate,3) 
	)
	,
	RMfileoftheMonth AS
	(
		Select 
		   [RELA CODE] = NULL
		  ,[VALIDATION CODE] = NULL
		  ,[BANK  CODE] = NULL
		  ,[BPI-AIA  BRANCH CODE]
		  ,[BRANCH CODE  FOR BPI-AIA]
		  ,[CLUSTER CODE] = NULL
		  ,[BANK]
		  ,[BRANCH NAME]
		  ,[ R BANK DIVISION] AS [DIVISION]
		  ,[R BANK AREA] AS [AREA]
		  ,[BRANCH TYPE]
		  ,[BUSINESS  MANAGER CODE] = NULL
		  ,[BUSINESS MANAGER] = NULL
		  ,[ASST BUSINESS MANAGER] = NULL
		  ,[BUSINESS DIRECTOR] 
		  ,[BPI DIVISION HEAD]
		  ,[RM REF CODE] = NULL
		  ,[RM] = NULL
		  ,[SA] = NULL
		  ,[SO] = NULL
		  ,[BRANCH_ADD] = NULL
		  ,[BRANCH_TEL] = NULL
		  ,[TEL 2] = NULL
		  ,[TEL 3] = NULL
		  ,[BRANCH_FAX] = NULL
		  ,[ZIP] = NULL
		  ,[TERRITORY  CODE]
		  ,[TERRITORY]
		  ,[BPI-AIA  DIVISION CODE]
		  ,[BPI-AIA DIVISION]
		  ,[BPI-AIA  AREA CODE]
		  ,[BPI-AIA AREA]
		  ,[BSE 1] AS [PERSONAL BSE 1]
		  ,[BSE1 CODE] AS [PER1 CODE]
		  ,[BSE 2] AS [PERSONAL BSE 2]
		  ,[BSE2 CODE] AS [PER2 CODE]
		  ,[BSE 3] AS [PERSONAL BSE 3]
		  ,[PER3 CODE] = NULL
		  ,[PERSONAL BSE 4] = NULL
		  ,[PER4 CODE] = NULL
		  ,[PERSONAL BSE 5] = NULL
		  ,[PER5 CODE] = NULL
		  ,[PREFERRED BSE 1] = NULL
		  ,[PRF1 CODE] = NULL
		  ,[PREFERRED BSE 2] = NULL
		  ,[PRF2 CODE] = NULL
		  ,[PREFERRED BSE 3] = NULL
		  ,[PRF3 CODE] = NULL
		  ,[BAM]
		  ,[BAM CODE]
		  ,[BDM]
		  ,[BDM CODE]
		  ,[TSH] = NULL
		  ,[TSH CODE] = NULL
		  ,[BRANCH OLD NAME] = NULL
		  ,[STATUS OF BRANCHES BASED ON BSC HIERARCHY] = NULL
		  ,[WWDate]
		,mlu.[Index]  FROM db_MasterFile_2024.[dbo].tbl_RBankMasterFile_2024 AS RMfile24 
		LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(RMfile24.WWDate,3) 
		) 

    SELECT MFile.* FROM MfileoftheMonth AS MFile WHERE MFile.[INDEX] = (SELECT MAX(mlu.[INDEX]) FROM MfileoftheMonth mlu)
	UNION ALL
	SELECT RMFile.* FROM RMfileoftheMonth AS RMFile WHERE RMFile.[INDEX] = (SELECT MAX(mlu.[INDEX]) FROM RMfileoftheMonth mlu)
'''
db_masterfile = sql_connection(query, 'db_MasterFile_2024')
db_masterfile

# %%
db_masterfile.columns

# %% [markdown]
# ## Merge BSE database to YTDN 

# %%
# Look Up BSE Code in YTDN to BSE Code in Master File
# Merged ytdn_2 and db_bse_masterfile into df_merge
ytdn_2 = ytdn_1.copy()

temp_df = db_bse_masterfile.drop_duplicates(subset=['BSE CODE']) # Remove duplicates so lookup merge only returns first match
db_bse_masterfile_tmp = temp_df.drop(['DEPLOYMENT DATE (MM/DD/YYYY)', 'TL NAME', 'JUNIOR/ SENIOR', 'TENURE  (CURRENT POSITION)', 'FIRST NAME',  'MOBILE NUMBERS', 'DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)', 'HYBRID', 'SURNAME', 'TSH', 'BATCH', 'WWDate', 'TL CODE', 'EE CODE', 'MIDDLE NAME', 'OLD/NEW',  'SEGMENT', 'TSH CODE', 'EMAILADD', 'POSITION',  'BSE FULL NAME'], axis=1)
ytdn_2['BSE CODE'] = ytdn_2['BSE CODE'].str.strip()
df_merge = ytdn_2.merge(db_bse_masterfile_tmp, left_on=['BSE CODE'], right_on=['BSE CODE'], how='left', suffixes=["", '_db_bse_masterfile'])
#df_merge_via_bsecode = df_merge.copy()
ytdn_2 = df_merge

ytdn_2


# %% [markdown]
# ## Merge MasterFile Database to YTDN (Left Join Performed)

# %%
## Please note the join performed is like a lookup Branch Code in YTDN to Branch Code in MasterFile-
       # need to pull the foloowing: 
       #'BRANCH CODE  FOR BPI-AIA', 'BRANCH NAME',
       #'DIVISION', 'AREA', 'BRANCH TYPE', 
       #'TERRITORY', 'BPI-AIA  DIVISION CODE', 'BPI-AIA DIVISION',
       #'BPI-AIA  AREA CODE', 'BPI-AIA AREA' BAM, BAM CODE, BDM, BDM CODE

ytdn_3 = ytdn_2.copy()

temp_df = db_masterfile.drop_duplicates(subset=['BRANCH CODE  FOR BPI-AIA']) # Remove duplicates so lookup merge only returns first match
db_masterfile_tmp = temp_df.drop(['BRANCH OLD NAME', 'VALIDATION CODE', 'ASST BUSINESS MANAGER', 'TERRITORY  CODE', 'PER1 CODE', 'PRF2 CODE', 'SO', 'WWDate', 'RM REF CODE', 'PER4 CODE', 'BRANCH_ADD', 'PREFERRED BSE 1', 'Index', 'SA', 'RM', 'RELA CODE', 'TEL 3', 'BANK  CODE', 'STATUS OF BRANCHES BASED ON BSC HIERARCHY', 'PREFERRED BSE 3', 'BUSINESS MANAGER', 'BPI-AIA  AREA CODE', 'PERSONAL BSE 1', 'TSH CODE', 'PER5 CODE', 'BRANCH_TEL', 'PERSONAL BSE 2', 'PERSONAL BSE 4', 'BUSINESS  MANAGER CODE', 'PER2 CODE', 'BPI DIVISION HEAD', 'TSH', 'PERSONAL BSE 5', 'PREFERRED BSE 2',  'PER3 CODE', 'CLUSTER CODE', 'BRANCH_FAX', 'PERSONAL BSE 3', 'PRF1 CODE', 'TEL 2', 'BANK', 'BUSINESS DIRECTOR', 'BPI-AIA  BRANCH CODE', 'ZIP', 'BPI-AIA  DIVISION CODE', 'PRF3 CODE'], axis=1)
ytdn_3['BRANCH CODE'] = ytdn_3['BRANCH CODE'].str.strip()
df_merge = ytdn_3.merge(db_masterfile_tmp, left_on=['BRANCH CODE'], right_on=['BRANCH CODE  FOR BPI-AIA'], how='left', suffixes=['', '_db_masterfile'])
#df_merge_via_branchcode = df_merge.copy()
ytdn_3 = df_merge.copy()

# Double Checking - Fill NaN Value during lookup using Branch Code with Valuse Coming from BSE Code in BSE Table

df_merge['BPI-AIA AREA'] = np.where(df_merge['BPI-AIA AREA'].isna(), df_merge['BPI-AIA AREA_db_masterfile'], df_merge['BPI-AIA AREA'])

df_merge['BPI-AIA DIVISION'] = np.where(df_merge['BPI-AIA DIVISION'].isna(), df_merge['BPI-AIA DIVISION_db_masterfile'], df_merge['BPI-AIA DIVISION'])

df_merge['BAM'] = np.where(df_merge['BAM'].isna(), df_merge['BAM_db_masterfile'], df_merge['BAM'])

df_merge['BAM CODE'] = np.where(df_merge['BAM CODE'].isna(), df_merge['BAM CODE_db_masterfile'], df_merge['BAM CODE'])

df_merge['BDM'] = np.where(df_merge['BDM'].isna(), df_merge['BDM_db_masterfile'], df_merge['BDM'])

df_merge['BDM CODE'] = np.where(df_merge['BDM CODE'].isna(), df_merge['BDM CODE_db_masterfile'], df_merge['BDM CODE'])

df_merge['TERRITORY'] = np.where(df_merge['TERRITORY'].isna(), df_merge['TERRITORY_db_masterfile'], df_merge['TERRITORY'])


## Setting Up Territory for DS and PB
df_merge['TERRITORY'] = np.where(df_merge['BPI-AIA DIVISION']=='DS', 'DS', df_merge['TERRITORY'])
df_merge['TERRITORY'] = np.where(df_merge['BPI-AIA DIVISION']=='PB', 'SG', df_merge['TERRITORY'])


## Putting default value to Bank Division, Bank Area, BPI AIA Division, BPI AIA Area, BDM Code, BDM, BAM Code, BAM

df_merge['BRANCH CODE'] = np.where(df_merge['BRANCH CODE'].isna(), 'SG', df_merge['BRANCH CODE'])

df_merge['AREA'] = np.where(df_merge['AREA'].isna(), 'SG', df_merge['AREA'])

df_merge['DIVISION'] = np.where(df_merge['DIVISION'].isna(), 'SG', df_merge['DIVISION'])

df_merge['BPI-AIA AREA'] = np.where(df_merge['BPI-AIA AREA'].isna(), 'SG', df_merge['BPI-AIA AREA'])

df_merge['BPI-AIA DIVISION'] = np.where(df_merge['BPI-AIA DIVISION'].isna(), 'SG', df_merge['BPI-AIA DIVISION'])

df_merge['BAM'] = np.where(df_merge['BAM'].isna(), 'SG', df_merge['BAM'])

df_merge['BAM CODE'] = np.where(df_merge['BAM CODE'].isna(), 'SG', df_merge['BAM CODE'])

df_merge['BDM'] = np.where(df_merge['BDM'].isna(), 'SG', df_merge['BDM'])

df_merge['BDM CODE'] = np.where(df_merge['BDM CODE'].isna(), 'SG', df_merge['BDM CODE'])

df_merge['TERRITORY'] = np.where(df_merge['TERRITORY'].isna(), 'SG', df_merge['TERRITORY'])



df_merge = df_merge.drop(['BPI-AIA AREA_db_masterfile', 'BPI-AIA DIVISION_db_masterfile', 'BAM_db_masterfile', 'BAM CODE_db_masterfile', 'BDM_db_masterfile', 'BDM CODE_db_masterfile'], axis=1)


ytdn_4 = df_merge.copy()


ytdn_4

# %% [markdown]
# ## Load Equivalent Mega Product Name for Plan Name

# %%
query = '''
    SELECT DISTINCT([PLAN NAME]) AS [PLAN NAME]
        ,[MEGA PRODUCT NAME]
        ,[MEGA PRODUCT NAME 2]
    FROM [db_MTD_Ref_2024].[dbo].[tbl_MTDRef_ProdName]
'''
db_mtdref_prod = sql_connection(query, 'db_MTD_Ref_2024')

# %% [markdown]
# ## Merge Mega Product Name with YTDN

# %%
# Using YTDN CAMPAIGN CODE to Match Plan Name to get Product Name

ytdn_5 = ytdn_4.copy()

# Merged df_merge_3 and db_mtdref_prod into df_merge
ytdn_5['CAMPAIGN CODE'] = ytdn_5['CAMPAIGN CODE'].str.strip()
df_merge = ytdn_5.merge(db_mtdref_prod, left_on=['CAMPAIGN CODE'], right_on=['PLAN NAME'], how='left', suffixes=['', '_db_mtdref_prod'])

# Deleted columns PLAN NAME_db_mtdref_prod, MEGA PRODUCT NAME 2
df_merge.drop(['PLAN NAME_db_mtdref_prod', 'MEGA PRODUCT NAME 2'], axis=1, inplace=True)

ytdn_5 = df_merge.copy()

ytdn_5

# %% [markdown]
# ## Merge Mega Product Name with YTDN 2

# %%
# Using YTDN PLAN NAME to Match Database PLAN NAME to get Product Name

ytdn_6 = ytdn_5.copy()

# Merged df_merge_3 and db_mtdref_prod into df_merge
ytdn_6['PLAN NAME'] = ytdn_6['PLAN NAME'].str.strip()
df_merge = ytdn_6.merge(db_mtdref_prod, left_on=['PLAN NAME'], right_on=['PLAN NAME'], how='left', suffixes=['', '_db_mtdref_prod'])

# Fill all the NaN with values from Product coming from matching Plan Names,
df_merge['MEGA PRODUCT NAME'] = np.where(df_merge['MEGA PRODUCT NAME'].isna(), df_merge['MEGA PRODUCT NAME_db_mtdref_prod'], df_merge['MEGA PRODUCT NAME'] )

# Deleted columns PLAN NAME_db_mtdref_prod, MEGA PRODUCT NAME 2
df_merge.drop(['MEGA PRODUCT NAME_db_mtdref_prod','MEGA PRODUCT NAME 2'], axis=1, inplace=True)

ytdn_6 = df_merge.copy()

print("Product Name Missing Values: ", ytdn_6['MEGA PRODUCT NAME'].isna().sum())
ytdn_6


# %%
if(len(ytdn_6.loc[ytdn_6['MEGA PRODUCT NAME'].isna()]) >0):
    ytdn_6.loc[ytdn_6['MEGA PRODUCT NAME'].isna()]
    raise  Exception("Check Missing Product Name")
else:
    pass

# %%
ytdn_6.loc[ytdn_6['MEGA PRODUCT NAME'].isna()]

# %% [markdown]
# ## ANP 100 

# %%
ytdn_7 = ytdn_6.copy()

# Added column 'ANP 100'
# ytdn_7['ANP 100'] = 0
# ytdn_7['ANP 100'] = np.where(ytdn_7['CA_CO']=='6051', (ytdn_7['ANNUAL PREMIUM'] * Dollar_Value), ytdn_7['ANNUAL PREMIUM'])

ytdn_7['ANP 100'] = 0

ytdn_7.loc[(ytdn_7['BILL MODE'] == '12') & (ytdn_7['CA_CO'] == '6051'), 'ANP 100'] = (ytdn_7['ANNUAL PREMIUM'] * Dollar_Value * 1)
ytdn_7.loc[(ytdn_7['BILL MODE'] == '12') & (ytdn_7['CA_CO'] == '6050'), 'ANP 100'] = (ytdn_7['ANNUAL PREMIUM'] * 1)

ytdn_7.loc[(ytdn_7['BILL MODE'] == '6') & (ytdn_7['CA_CO'] == '6051'), 'ANP 100'] = (ytdn_7['SEMIANNUAL PREMIUM'] * Dollar_Value * 2)
ytdn_7.loc[(ytdn_7['BILL MODE'] == '6') & (ytdn_7['CA_CO'] == '6050'), 'ANP 100'] = (ytdn_7['SEMIANNUAL PREMIUM'] * 2)

ytdn_7.loc[(ytdn_7['BILL MODE'] == '3') & (ytdn_7['CA_CO'] == '6051'), 'ANP 100'] = (ytdn_7['QUARTERLY PREMIUM'] * Dollar_Value * 4)
ytdn_7.loc[(ytdn_7['BILL MODE'] == '3') & (ytdn_7['CA_CO'] == '6050'), 'ANP 100'] = (ytdn_7['QUARTERLY PREMIUM'] * 4)

ytdn_7.loc[(ytdn_7['BILL MODE'] == '1') & (ytdn_7['CA_CO'] == '6051'), 'ANP 100'] = (ytdn_7['MONTHLY PREMIUM'] * Dollar_Value * 12)
ytdn_7.loc[(ytdn_7['BILL MODE'] == '1') & (ytdn_7['CA_CO'] == '6050'), 'ANP 100'] = (ytdn_7['MONTHLY PREMIUM'] * 12)

ytdn_7.loc[ytdn_7['PLAN NAME'] == 'RP DOLLAR 10-PAY BAND 1'].head()


# %%
ytdn_7.head()

# %%
ytdn_7.loc[ytdn_7['CA_POLICY NO.'] == '7013115750']

# %% [markdown]
# ## MEGA RP/SP Product Name

# %%
ytdn_8 = ytdn_7.copy()

# Added column new-column-a1a8
ytdn_8['RP/SP'] = 0

# Changed new-column-a1a8 to dtype str
ytdn_8['RP/SP'] = ytdn_8['RP/SP'].astype('str')

df_ytdn8_1 = ytdn_8.copy()

sp_products = ["INVEST PLUS DOLLAR", "INVEST DOLLAR MAX", "INVEST PLUS PESO", "INVEST PESO MAX", "PREMIER BENEFIT LIFE (1-Pay)", "PREFERRED LIFE PLUS", "INCOME BOOSTER 5"]

#df_ytdn2['RP/SP'] = np.where(df_ytdn2['MEGA PRODUCT NAME'].isin(sp_products))


## Single Pay | Regular Pay
df_ytdn8_1['RP/SP'] = np.where(df_ytdn8_1['MEGA PRODUCT NAME'].isin(sp_products), "SP", "RP")

ytdn_8 = df_ytdn8_1.copy()

del(df_ytdn8_1)

ytdn_8.head()



# %% [markdown]
# ## ANP in Pesos and Client Segment 2

# %%
ytdn_9 = ytdn_8.copy()

# Added column ANP in PESOS
ytdn_9.insert(39, 'ANP in PESOS', 0)

ytdn_9['ANP in PESOS'] = np.where((ytdn_9['RP/SP'] == "SP"), (ytdn_9['ANP 100'] * (10/100)), (ytdn_9['ANP 100']))
ytdn_9['ANP in PESOS'] = np.where(((ytdn_9['MEGA PRODUCT NAME'].str.upper() == "Dollar Protect Plus".upper()) & (ytdn_9['CA_CO'] != '6051')), (ytdn_9["ANP 100"]) * Dollar_Value, (ytdn_9['ANP in PESOS']))
ytdn_9['ANP in PESOS'] = np.where(((ytdn_9['MEGA PRODUCT NAME'].str.upper() == "Vitality Dollar Protect Plus".upper()) & (ytdn_9['CA_CO'] != '6051')), (ytdn_9["ANP 100"]) * Dollar_Value, (ytdn_9['ANP in PESOS']))

# Added column CLIENT SEGMENT 2
ytdn_9['CLIENT SEGMENT 2'] = 0
ytdn_9['CLIENT SEGMENT 2'] = ytdn_9['CLIENT SEGMENT 2'].astype('str')

conditions  = [ ytdn_9['MARKET SEGMENT'] == 'I01', ytdn_9['MARKET SEGMENT'] == 'I02', ytdn_9['MARKET SEGMENT'] == 'I03', ytdn_9['MARKET SEGMENT'] == 'I04', ytdn_9['MARKET SEGMENT'] == 'I05',  ytdn_9['MARKET SEGMENT'] == 'I06',  ytdn_9['MARKET SEGMENT'] == 'I07']
choices     = [ "Private", "Preferred", "Personal", "Corporate", "OF", "Corporate", "Corporate" ]

# df_ytdn3['CLIENT SEGMENT 2'] = np.select(conditions, choices, default=np.nan)
ytdn_9['CLIENT SEGMENT 2'] = np.select(conditions, choices, default="Preferred")

ytdn_9.head()


# %% [markdown]
# ## Extract Date

# %%
ytdn_10 = ytdn_9.copy()

# Added column EXTRACT DATE
ytdn_10['EXTRACT DATE'] = 0

# Changed EXTRACT DATE to dtype datetime
ytdn_10['EXTRACT DATE'] = pd.to_datetime(ytdn_10['EXTRACT DATE'], unit='s', errors='coerce')

# Add EXTRACT DATE to dtype datetime
ytdn_10['EXTRACT DATE'] = datetime.today()

ytdn_10.head()

# %% [markdown]
# ## Merge for OLD/NEW from BSE Database

# %%
ytdn_11 = ytdn_10.copy()

# Merged df_ytdn4 and db_bse_masterfile into df_merge
temp_df = db_bse_masterfile.drop_duplicates(subset=['BSE CODE']) # Remove duplicates so lookup merge only returns first match
db_bse_masterfile_tmp = temp_df.drop(['BSE FULL NAME', 'MIDDLE NAME', 'TSH', 'BPI-AIA AREA', 'BPI-AIA DIVISION', 'JUNIOR/ SENIOR', 'TL NAME',  'POSITION', 'HYBRID', 'EMAILADD', 'TL CODE', 'BATCH', 'EE CODE', 'MOBILE NUMBERS', 'TSH CODE', 'SURNAME', 'DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)', 'BDM', 'BAM', 'BDM CODE', 'TENURE  (CURRENT POSITION)', 'DEPLOYMENT DATE (MM/DD/YYYY)', 'TERRITORY', 'FIRST NAME', 'WWDate', 'BAM CODE'], axis=1)
df_merge = ytdn_11.merge(db_bse_masterfile_tmp, left_on=['BSE CODE'], right_on=['BSE CODE'], how='left', suffixes=['', '_db_bse_masterfile'])
ytdn_11 =  df_merge.copy()

# Filling OLD/NEW
# ytdn_11['OLD/NEW'] = np.where((ytdn_11['OLD/NEW'].isna()), "NEW", ytdn_11['OLD/NEW']) #Tag as Old
ytdn_11['OLD/NEW'] = np.where((ytdn_11['OLD/NEW'].isna()), "NEW", ytdn_11['OLD/NEW']) #Tag as Old


# Filling SEGMENT
ytdn_11['SEGMENT'] = np.where((ytdn_11['SEGMENT'].isna()), "Not in Masterlist", ytdn_11['SEGMENT'])


# Renamed columns BSE SEGMENT
ytdn_11.rename(columns={'SEGMENT': 'BSE SEGMENT'}, inplace=True)

# Added column Ok/Not Ok
ytdn_11.insert(43, 'Ok/Not Ok', 0)

# Changed Ok/Not Ok to dtype str
ytdn_11['Ok/Not Ok'] = ytdn_11['Ok/Not Ok'].astype('str')

ytdn_11['Ok/Not Ok'] = np.where(ytdn_11['BSE SEGMENT'] == "PRIVATE", "Not Ok", "Ok")

ytdn_11.head()


# %% [markdown]
# ## Load Referrer List Database

# %%
query = '''

WITH Tbl_Bank_RM AS
(
	SELECT 
		DISTINCT(TRIM(Bank24.[REFR CODE])) AS [Referrer Code]
		,Bank24.[FULL NAME] AS [Full Name]
		,Bank24.[FUNCTION] AS [Position]
		,Bank24.[WWDate]
	FROM [db_MasterFile_2024].[dbo].[tbl_BankEmployee_2024] as Bank24

	UNION ALL

	SELECT 
		DISTINCT(TRIM(RBank24.[REFR CODE])) AS [Referrer Code]
		,RBank24.[FULL NAME] AS [Full Name]
		,RBank24.[FUNCTION] AS [Position]
		,RBank24.[WWDate]
	FROM [db_MasterFile_2024].[dbo].[tbl_RBankEmployee_2024] as RBank24

	UNION ALL

	SELECT 
		 DISTINCT(TRIM(RM24.[REFR CODE])) AS [Referrer Code]
		 ,RM24.[RM] AS [Full Name]
		 ,[Position] = 'RM'
		 ,[WWDate]
	 FROM [db_MasterFile_2024].[dbo].[tbl_RM_BPIAIA_Area_2024] AS RM24
),
Tbl_Bank_RM_Current AS
(
	SELECT 	 
		DISTINCT(TRIM(Tbl_Bank_RM.[Referrer Code])) AS [Referrer Code]
		,Tbl_Bank_RM.[Full Name] AS [Full Name]
		,Tbl_Bank_RM.[Position] AS [Position]
		,Tbl_Bank_RM.[WWDate]
		,miu.[INDEX]
	 FROM Tbl_Bank_RM 
	 LEFT JOIN db_Lookup.[dbo].tbl_monthindex_unique as miu 
	 ON miu.[MON] = LEFT(Tbl_Bank_RM.[WWDate],3)
)

SELECT 
	DISTINCT(TRIM(TRBC.[Referrer Code])) AS [Referrer Code]
	,TRBC.[Full Name]
	,TRBC.[Position]
	FROM Tbl_Bank_RM_Current AS TRBC WHERE [INDEX] = 
	(SELECT MAX(Tbl_Bank_RM_Current.[INDEX]) FROM Tbl_Bank_RM_Current)

'''
db_ref_list = sql_connection(query, 'db_MasterFile_2024')

# %% [markdown]
# ## Merge Referrer List Database

# %%
## REFNAME, REFCAT2, REFCAT3

# Match ytdn Referrer Code to  database tbl_Referrer_List Referrer Code

# Merged ytdn and db_ref_list 
ytdn_12 = ytdn_11.copy()

temp_df = db_ref_list.drop_duplicates(subset=['Referrer Code']) # Remove duplicates so lookup merge only returns first match
ytdn_12['REFERRER CODE'] = ytdn_12['REFERRER CODE'].str.strip()
ytdn_12 = ytdn_12.merge(temp_df, left_on=['REFERRER CODE'], right_on=['Referrer Code'], how='left', suffixes=['', '_db_ref_list'])

# Renamed columns REFNAME, REFCAT2
ytdn_12.rename(columns={'Full Name': 'REFNAME', 'Position': 'REFCAT2'}, inplace=True)

ytdn_12['REFNAME'] = np.where(ytdn_12['REFNAME'].isna(), "Not found in DB", ytdn_12['REFNAME'])
ytdn_12['REFCAT2'] = np.where(ytdn_12['REFCAT2'].isna(), "Not found in DB", ytdn_12['REFCAT2'])

ytdn_12['BRANCH TYPE'] = np.where(ytdn_12['BRANCH TYPE'].isna(), "blank", ytdn_12['BRANCH TYPE'])

# Added column REFCAT3
ytdn_12.insert(50, 'REFCAT3', 0)

ytdn_12['REFCAT3'] = np.where((ytdn_12['REFCAT2'] == "CRS") | (ytdn_12['REFCAT2'] == "CSSA"), "BRANCH", ytdn_12['REFCAT2'])

ytdn_12.head()


# %% [markdown]
# ## DPP finder, Pay Variant Finder, REG/Mat Recap, CC100 finder, Pay Variant Finder, TCM finder, Pay Variant Finder

# %%
ytdn_13 = ytdn_12.copy()

# Added column DPP Finder
ytdn_13['DPP Finder'] = 0
ytdn_13['DPP Finder'] = np.where(ytdn_13['MEGA PRODUCT NAME'].str.upper().str.contains("DOLLAR PROTECT"), ytdn_13['MEGA PRODUCT NAME'], "-")

# Added column Pay Variant Finder
ytdn_13['Pay Variant Finder'] = 0
ytdn_13['Pay Variant Finder'] = np.where((ytdn_13['PLAN NAME'].str.upper().str.contains("REGULAR PAY")) | (ytdn_13['PLAN NAME'].str.upper().str.contains("REGULAR-PAY")), "REGULAR PAY", "-")
ytdn_13['Pay Variant Finder'] = np.where((ytdn_13['PLAN NAME'].str.upper().str.contains("10-PAY")) | (ytdn_13['PLAN NAME'].str.upper().str.contains("10 PAY")), "10-Pay", ytdn_13['Pay Variant Finder'])

# Added column REG/Mat Recap
# Table for D Series Plans
D_Plans = {
    "Plan Name" : ["D1904BO", "D2103BO"],
    "Product Name": ["VIP DPP Reg", "DPP Mat Recap"]
}

D_Plans = pd.DataFrame.from_dict(D_Plans)


temp_df = D_Plans.drop_duplicates(subset=['Plan Name']) # Remove duplicates so lookup merge only returns first match
ytdn_13['CAMPAIGN CODE'] = ytdn_13['CAMPAIGN CODE'].str.strip()
df_merge = ytdn_13.merge(temp_df, left_on=['CAMPAIGN CODE'], right_on=['Plan Name'], how='left', suffixes=['ytdn_13', '_D_Plans'])

# Deleted columns Plan Name
df_merge.drop(['Plan Name'], axis=1, inplace=True)

# Rename to REG/Mat Recap
df_merge.rename(columns={'Product Name': 'REG/Mat Recap'}, inplace=True)
df_merge['REG/Mat Recap'] = np.where(df_merge['REG/Mat Recap'].isna(), "-", df_merge['REG/Mat Recap'])

ytdn_13 = df_merge.copy()

# Added column CC100 finder
ytdn_13['CC100 finder'] = 0
ytdn_13['CC100 finder'] = np.where(ytdn_13['MEGA PRODUCT NAME'].str.upper().str.contains("CRITICAL CARE 100"), ytdn_13['MEGA PRODUCT NAME'], "-")

# Added column Pay Variant Finder 2
ytdn_13['Pay Variant Finder 2'] = 0
ytdn_13['Pay Variant Finder 2'] = np.where((ytdn_13['PLAN NAME'].str.upper().str.contains("20")), "20-Pay", "-")
ytdn_13['Pay Variant Finder 2'] = np.where((ytdn_13['PLAN NAME'].str.upper().str.contains("10")), "10-Pay", ytdn_13['Pay Variant Finder 2'])

# Added column TCM finder
ytdn_13['TCM finder'] = 0
ytdn_13['TCM finder'] = np.where((ytdn_13['MEGA PRODUCT NAME'].str.upper().str.contains("TOTAL CARE MAX")), ytdn_13['MEGA PRODUCT NAME'], "-")

# Added column Pay Variant Finder 3
ytdn_13['Pay Variant Finder 3'] = 0
ytdn_13['Pay Variant Finder 3'] = np.where((ytdn_13['PLAN NAME'].str.upper().str.contains("20")), "20-Pay", "-")
ytdn_13['Pay Variant Finder 3'] = np.where((ytdn_13['PLAN NAME'].str.upper().str.contains("10")), "10-Pay", ytdn_13['Pay Variant Finder 3'])
ytdn_13['Pay Variant Finder 3'] = np.where((ytdn_13['PLAN NAME'].str.upper().str.contains("65")), "Pay-to-age 65", ytdn_13['Pay Variant Finder 3'])

ytdn_13.head()

# %% [markdown]
# ## BPL Monitoring

# %% [markdown]
# ### Merge db_mtdref_prod to ytdn via PLAN NAME

# %%
ytdn_14 = ytdn_13.copy()

# Merged df_ytdn8 and db_mtdref_prod into df_ytdn_14
temp_df = db_mtdref_prod.drop_duplicates(subset=['PLAN NAME']) # Remove duplicates so lookup merge only returns first match
db_mtdref_prod_tmp = temp_df.drop(['MEGA PRODUCT NAME'], axis=1)
df_merge = ytdn_14.merge(db_mtdref_prod_tmp, left_on=['PLAN NAME'], right_on=['PLAN NAME'], how='left', suffixes=['_ytdn14', '_db_mtdref_prod'])

ytdn_14 = df_merge.copy()
ytdn_14.head()



# %% [markdown]
# ### Merge db_mtdref_prod to ytdn via PLAN NAME and Campaign Code

# %%
# Merged df_ytdn8 and db_mtdref_prod into df_merge
ytdn_14_2 = ytdn_14.copy()
temp_df = db_mtdref_prod.drop_duplicates(subset=['PLAN NAME']) # Remove duplicates so lookup merge only returns first match
db_mtdref_prod_tmp = temp_df.drop(['MEGA PRODUCT NAME'], axis=1)
df_merge = ytdn_14_2.merge(db_mtdref_prod_tmp, left_on=['CAMPAIGN CODE'], right_on=['PLAN NAME'], how='left', suffixes=['', '_db_mtdref_prod'])

# Deleted columns PLAN NAME_db_mtdref_prod
#df_merge.drop(['PLAN NAME_db_mtdref_prod'], axis=1, inplace=True)

ytdn_14_2 = df_merge.copy()

ytdn_14_2.head()

# %% [markdown]
# ### BLP Column Adding

# %%
# Added column BLP MONITORING
ytdn_14_3 = ytdn_14_2.copy()

ytdn_14_3['BLP MONITORING'] = '-'
ytdn_14_3['BLP MONITORING'] = np.where(ytdn_14_3['MEGA PRODUCT NAME 2'].isna(), ytdn_14_3['BLP MONITORING'], ytdn_14_3['MEGA PRODUCT NAME 2'])
ytdn_14_3['BLP MONITORING'] = np.where((ytdn_14_3['MEGA PRODUCT NAME 2_db_mtdref_prod'].notna()) & (ytdn_14_3['BLP MONITORING'] == '-'),  ytdn_14_3['MEGA PRODUCT NAME 2_db_mtdref_prod'], ytdn_14_3['BLP MONITORING'])

# Deleted columns MEGA PRODUCT NAME 2_df_ytdn8, MEGA PRODUCT NAME 2_db_mtdref_prod
ytdn_14_3.drop(['MEGA PRODUCT NAME 2', 'PLAN NAME_db_mtdref_prod', 'MEGA PRODUCT NAME 2_db_mtdref_prod'], axis=1, inplace=True)

ytdn_14 = ytdn_14_3.copy()

ytdn_14.head()

# %% [markdown]
# ## DS AREA, DS BAM CODE, DS BAM NAME

# %%
ytdn_15 = ytdn_14.copy()

db_dsbse_masterfile_ = db_bse_masterfile[db_bse_masterfile['POSITION'].apply(lambda val: all(val != s for s in ['ABBSE', 'BSE', 'MBSE', 'MBSE ', 'PBIS']))].copy()

# Merged df_ytdn9 and db_dsbse_masterfile_ into df_merge
temp_df = db_dsbse_masterfile_.drop_duplicates(subset=['BSE CODE']) # Remove duplicates so lookup merge only returns first match
db_dsbse_masterfile__tmp = temp_df.drop(['JUNIOR/ SENIOR', 'TL NAME', 'DEPLOYMENT DATE (MM/DD/YYYY)', 'BPI-AIA DIVISION',  'BSE FULL NAME', 'TSH CODE',  'SEGMENT', 'TENURE  (CURRENT POSITION)', 'SURNAME', 'EMAILADD', 'BATCH', 'TL CODE',  'TERRITORY', 'DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)', 'TSH', 'BDM', 'WWDate', 'MIDDLE NAME', 'HYBRID', 'MOBILE NUMBERS', 'POSITION', 'EE CODE', 'OLD/NEW', 'BDM CODE', 'FIRST NAME'], axis=1)
df_merge = ytdn_15.merge(db_dsbse_masterfile__tmp, left_on=['BSE CODE'], right_on=['BSE CODE'], how='left', suffixes=['', '_db_dsbse_masterfile_'])

ytdn_15 =df_merge.copy()

# Renamed columns DS AREA, DS BAM CODE, DS BAM NAME
ytdn_15.rename(columns={'BPI-AIA AREA_db_dsbse_masterfile_': 'DS AREA', 'BAM CODE_db_dsbse_masterfile_': 'DS BAM CODE', 'BAM_db_dsbse_masterfile_': 'DS BAM NAME'}, inplace=True)

ytdn_15['DS AREA'] = np.where(ytdn_15['DS AREA'].isna(), "-", ytdn_15['DS AREA'])
ytdn_15['DS BAM CODE'] = np.where(ytdn_15['DS BAM CODE'].isna(), "-", ytdn_15['DS BAM CODE'])
ytdn_15['DS BAM NAME'] = np.where(ytdn_15['DS BAM NAME'].isna(), "-", ytdn_15['DS BAM NAME'])


ytdn_15.head()

# %% [markdown]
# ## Load EKYFC Database

# %%
query = '''
SELECT TRIM([CompanyCode]) AS [CompanyCode]
      ,[agentcode1]
      ,[PolicyNo]
      ,[SubmissioNDateTime]
      ,[iPOSVersion]
      ,[RemoteSignature]
      ,[RemoteEPayment]
   FROM [db_DSR].[dbo].[EKYC]

'''
db_ekyfc_list = sql_connection(query, 'db_DSR')

# %% [markdown]
# ## EKYFC Policy No Merge to YTDN CA_POLICY_No

# %%
ytdn_16 = ytdn_15.copy()


# Merged df_ytdn10 and db_ekyfc_list into df_merge
temp_df = db_ekyfc_list.drop_duplicates(subset=['PolicyNo']) # Remove duplicates so lookup merge only returns first match
db_ekyfc_list_tmp = temp_df.drop(['iPOSVersion', 'agentcode1', 'SubmissioNDateTime', 'CompanyCode'], axis=1)
ytdn_16['CA_POLICY NO.'] = ytdn_16['CA_POLICY NO.'].str.strip()
df_merge = ytdn_16.merge(db_ekyfc_list_tmp, left_on=['CA_POLICY NO.'], right_on=['PolicyNo'], how='left', suffixes=['_df_ytdn10', '_db_ekyfc_list'])

ytdn_16 = df_merge.copy()

ytdn_16['RemoteSignature'] = np.where(ytdn_16['RemoteSignature'].isna(), "Not in eKYC list", ytdn_16['RemoteSignature'])
ytdn_16['RemoteEPayment'] = np.where(ytdn_16['RemoteEPayment'].isna(), "Not in eKYC list", ytdn_16['RemoteEPayment'])


# Renamed columns eKYC_RemoteSignature, eKYC_RemoteEPayment
ytdn_16.rename(columns={'RemoteSignature': 'eKYC_RemoteSignature', 'RemoteEPayment': 'eKYC_RemoteEPayment'}, inplace=True)

ytdn_16.head()


# %% [markdown]
# ## Load MyData Database

# %%
query = '''
SELECT TRIM([PolicyNo]) AS [PolicyNo]
      ,[DDUW_Ind]
	  ,[Remarks] = 'Remarks'
  FROM [db_DSR].[dbo].[Mydata_Pol]
'''
db_mydata = sql_connection(query, 'db_DSR')


# %% [markdown]
# ## MyData Transaformation

# %%
ytdn_17 = ytdn_16.copy()

ytdn_17['PolicyNo'] = ytdn_17['PolicyNo'].str.strip()
# Merged df_ytdn11 and db_mydata into df_merge
temp_df = db_mydata.drop_duplicates(subset=['PolicyNo']) # Remove duplicates so lookup merge only returns first match
db_mydata_tmp = temp_df.drop(['Remarks'], axis=1)
ytdn_17['PolicyNo'] = ytdn_17['PolicyNo'].str.strip()
df_merge = ytdn_17.merge(db_mydata_tmp, left_on=['PolicyNo'], right_on=['PolicyNo'], how='left', suffixes=['_ytdn_17', '_db_mydata'])

# Renamed columns MyData
df_merge.rename(columns={'DDUW_Ind': 'MyData'}, inplace=True)

ytdn_17 = df_merge.copy()

ytdn_17['MyData'] = np.where(ytdn_17['MyData'].isna(), "Not in MyData List", ytdn_17['MyData'])

ytdn_17.head()

# %% [markdown]
# ## TCM Monitoring

# %%
ytdn_18 = ytdn_17.copy()

ytdn_18['PLAN NAME'] = ytdn_18['PLAN NAME'].str.strip()

db_mtdref_prod_totalcaremax = db_mtdref_prod.loc[(db_mtdref_prod['PLAN NAME'] == 'Total Care Max 10-PAY') | (db_mtdref_prod['PLAN NAME'] == 'Total Care Max 20-PAY')]

# Merged df_ytdn12 and db_mtdref_prod_totalcaremax into df_merge via 
temp_df = db_mtdref_prod_totalcaremax.drop_duplicates(subset=['PLAN NAME']) # Remove duplicates so lookup merge only returns first match
db_mtdref_prod_totalcaremax_tmp = temp_df.drop(['MEGA PRODUCT NAME'], axis=1)
df_merge = ytdn_18.merge(db_mtdref_prod_totalcaremax_tmp, left_on=['PLAN NAME'], right_on=['PLAN NAME'], how='left', suffixes=['', '_db_mtdref_prod_totalcaremax'])

# Renamed columns TCM MONITORING
df_merge.rename(columns={'MEGA PRODUCT NAME 2': 'TCM MONITORING'}, inplace=True)

ytdn_18 = df_merge.copy()
ytdn_18['CAMPAIGN CODE'] = ytdn_18['CAMPAIGN CODE'].str.strip()

# Merged df_ytdn12 and db_mtdref_prod_totalcaremax into df_merge
temp_df = db_mtdref_prod_totalcaremax.drop_duplicates(subset=['PLAN NAME']) # Remove duplicates so lookup merge only returns first match
db_mtdref_prod_totalcaremax_tmp = temp_df.drop(['MEGA PRODUCT NAME'], axis=1)
df_merge = ytdn_18.merge(db_mtdref_prod_totalcaremax_tmp, left_on=['CAMPAIGN CODE'], right_on=['PLAN NAME'], how='left', suffixes=['_df_ytdn12', '_db_mtdref_prod_totalcaremax'])

# Deleted columns PLAN NAME_db_mtdref_prod_totalcaremax
df_merge.drop(['PLAN NAME_db_mtdref_prod_totalcaremax'], axis=1, inplace=True)

# Renamed columns TCM MONITORING 2
df_merge.rename(columns={'MEGA PRODUCT NAME 2': 'TCM MONITORING 2'}, inplace=True)

ytdn_18 = df_merge.copy()

ytdn_18['TCM MONITORING'] = np.where(ytdn_18['TCM MONITORING'].isna(), ytdn_18['TCM MONITORING 2'], ytdn_18['TCM MONITORING'])
ytdn_18['TCM MONITORING'] = np.where(ytdn_18['TCM MONITORING'].isna(), "-", ytdn_18['TCM MONITORING'])

# Deleted columns 'TCM MONITORING 2'
ytdn_18.drop(['TCM MONITORING 2'], axis=1, inplace=True)

ytdn_18.head()


# %% [markdown]
# ## Arranging and Renaming Columns

# %%
# Renamed columns CA, CA_INAME, CA_BM, CA_APREM, CA_SPREM, CA_QPREM, CA_MPREM, CB_COVAMT, CB_PLAN, PN_PLAN_NAME

ytdn_19 = ytdn_18.copy()

ytdn_19 = ytdn_19[['CA_CO', 'SUBMISSION DATE', 'SUBMISSION TIME', 'CA_POLICY NO.',
       'CA_NAME', 'SUSAMT1', 'APPROVAL DATE', 'ISSUE DATE', 'STATUS CODE',
       'BILL MODE', 'ANNUAL PREMIUM', 'SEMIANNUAL PREMIUM',
       'QUARTERLY PREMIUM', 'MONTHLY PREMIUM', 'COVERAGE AMOUNT', 'PLAN CODE',
       'PLAN NAME_df_ytdn12', 'PHASE', 'SUBPHASE', 'BRANCH CODE', 'BSE CODE',
       'BSE NAME', 'REFERRER CODE', 'MARKET SEGMENT', 'CAMPAIGN CODE',
       'EXTRACTED DATE', 'BRANCH CODE  FOR BPI-AIA', 'BRANCH NAME', 'DIVISION',
       'AREA', 'BRANCH TYPE', 'TERRITORY', 'BPI-AIA DIVISION',
       'BPI-AIA AREA', 'BAM', 'BAM CODE', 'BDM',
       'ANP 100', 'RP/SP', 'ANP in PESOS', 'CLIENT SEGMENT 2', 'BDM CODE',
       'EXTRACT DATE', 'Ok/Not Ok', 'MEGA PRODUCT NAME', 'BSE SEGMENT',
       'OLD/NEW', 'Referrer Code', 'REFNAME', 'REFCAT2', 'REFCAT3',
       'DPP Finder', 'Pay Variant Finder', 'REG/Mat Recap', 'CC100 finder',
       'Pay Variant Finder 2', 'TCM finder', 'Pay Variant Finder 3',
       'BLP MONITORING', 'DS AREA',
       'DS BAM NAME', 'DS BAM CODE', 'PolicyNo', 'eKYC_RemoteSignature',
       'eKYC_RemoteEPayment', 'MyData', 'TCM MONITORING']]

ytdn_20 = ytdn_19.copy()

ytdn_20.rename(columns={
    'CA_CO': 'CA_CO',
    'CA_NAME': 'CA_INAME',
    'CA_POLICY NO.':'CA_POLNUM',
    'BILL MODE': 'CA_BM',
    'ANNUAL PREMIUM': 'CA_APREM',
    'SEMIANNUAL PREMIUM': 'CA_SPREM',
    'QUARTERLY PREMIUM': 'CA_QPREM',
    'MONTHLY PREMIUM': 'CA_MPREM',
    'COVERAGE AMOUNT': 'CB_COVAMT',
    'PLAN CODE': 'CB_PLAN',
    'PLAN NAME_df_ytdn12': 'PN_PLAN_NAME'
}, inplace=True)

# Reordered column OLD/NEW
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'OLD/NEW']
ytdn_20_columns.insert(18, 'OLD/NEW')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column BSE CODE
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BSE CODE']
ytdn_20_columns.insert(19, 'BSE CODE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns CC_SERVAGT
ytdn_20.rename(columns={'BSE CODE': 'CC_SERVAGT'}, inplace=True)

# Reordered column BSE NAME
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BSE NAME']
ytdn_20_columns.insert(20, 'BSE NAME')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns EA_AGTNAME
ytdn_20.rename(columns={'BSE NAME': 'EA_AGTNAME'}, inplace=True)

# Reordered column BRANCH CODE
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BRANCH CODE']
ytdn_20_columns.insert(21, 'BRANCH CODE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Deleted columns BRANCH CODE  FOR BPI-AIA
ytdn_20.drop(['BRANCH CODE  FOR BPI-AIA'], axis=1, inplace=True)

# Reordered column BRANCH NAME
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BRANCH NAME']
ytdn_20_columns.insert(22, 'BRANCH NAME')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column DIVISION
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'DIVISION']
ytdn_20_columns.insert(23, 'DIVISION')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns BANK DIVISION
ytdn_20.rename(columns={'DIVISION': 'BANK DIVISION'}, inplace=True)

# Reordered column AREA
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'AREA']
ytdn_20_columns.insert(24, 'AREA')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns BANK AREA
ytdn_20.rename(columns={'AREA': 'BANK AREA'}, inplace=True)

# Reordered column BPI-AIA DIVISION
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BPI-AIA DIVISION']
ytdn_20_columns.insert(25, 'BPI-AIA DIVISION')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column BPI-AIA AREA
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BPI-AIA AREA']
ytdn_20_columns.insert(26, 'BPI-AIA AREA')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns BPI AIA AREA, BPI AIA DIVISION
ytdn_20.rename(columns={'BPI-AIA AREA': 'BPI AIA AREA', 'BPI-AIA DIVISION': 'BPI AIA DIVISION'}, inplace=True)

# Reordered column BDM CODE
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BDM CODE']
ytdn_20_columns.insert(27, 'BDM CODE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column BDM
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BDM']
ytdn_20_columns.insert(27, 'BDM')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column BAM CODE_df_ytdn9
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BAM CODE']
ytdn_20_columns.insert(29, 'BAM CODE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns BAM CODE
ytdn_20.rename(columns={'BAM CODE': 'BAM CODE'}, inplace=True)

# Reordered column BAM_df_ytdn9
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BAM']
ytdn_20_columns.insert(30, 'BAM')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns BAM
ytdn_20.rename(columns={'BAM_df_ytdn9': 'BAM'}, inplace=True)

# Reordered column MEGA PRODUCT NAME
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'MEGA PRODUCT NAME']
ytdn_20_columns.insert(31, 'MEGA PRODUCT NAME')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns PRODUCT
ytdn_20.rename(columns={'MEGA PRODUCT NAME': 'PRODUCT'}, inplace=True)

# Reordered column ANP 100
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'ANP 100']
ytdn_20_columns.insert(32, 'ANP 100')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns ANP (SP @ 100%)
ytdn_20.rename(columns={'ANP 100': 'ANP (SP @ 100%)'}, inplace=True)

# Reordered column RP/SP
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'RP/SP']
ytdn_20_columns.insert(33, 'RP/SP')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column ANP in PESOS
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'ANP in PESOS']
ytdn_20_columns.insert(34, 'ANP in PESOS')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns ANP in PESO
ytdn_20.rename(columns={'ANP in PESOS': 'ANP in PESO'}, inplace=True)

# Reordered column REFERRER CODE
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'REFERRER CODE']
ytdn_20_columns.insert(35, 'REFERRER CODE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column MARKET SEGMENT
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'MARKET SEGMENT']
ytdn_20_columns.insert(36, 'MARKET SEGMENT')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Renamed columns CLIENT SEGMENT
ytdn_20.rename(columns={'MARKET SEGMENT': 'CLIENT SEGMENT'}, inplace=True)

# Reordered column CLIENT SEGMENT 2
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'CLIENT SEGMENT 2']
ytdn_20_columns.insert(37, 'CLIENT SEGMENT 2')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column EXTRACT DATE
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'EXTRACT DATE']
ytdn_20_columns.insert(38, 'EXTRACT DATE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column BSE SEGMENT
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BSE SEGMENT']
ytdn_20_columns.insert(39, 'BSE SEGMENT')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column Ok/Not Ok
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'Ok/Not Ok']
ytdn_20_columns.insert(40, 'Ok/Not Ok')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column CAMPAIGN CODE
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'CAMPAIGN CODE']
ytdn_20_columns.insert(41, 'CAMPAIGN CODE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column TERRITORY
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'TERRITORY']
ytdn_20_columns.insert(42, 'TERRITORY')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column DPP Finder
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'DPP Finder']
ytdn_20_columns.insert(43, 'DPP Finder')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column Pay Variant Finder
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'Pay Variant Finder']
ytdn_20_columns.insert(44, 'Pay Variant Finder')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column REG/Mat Recap
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'REG/Mat Recap']
ytdn_20_columns.insert(45, 'REG/Mat Recap')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column CC100 finder
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'CC100 finder']
ytdn_20_columns.insert(46, 'CC100 finder')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column Pay Variant Finder 2
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'Pay Variant Finder 2']
ytdn_20_columns.insert(47, 'Pay Variant Finder 2')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column TCM finder
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'TCM finder']
ytdn_20_columns.insert(48, 'TCM finder')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column Pay Variant Finder 3
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'Pay Variant Finder 3']
ytdn_20_columns.insert(49, 'Pay Variant Finder 3')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column REFNAME
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'REFNAME']
ytdn_20_columns.insert(50, 'REFNAME')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column REFCAT2
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'REFCAT2']
ytdn_20_columns.insert(51, 'REFCAT2')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column BRANCH TYPE
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BRANCH TYPE']
ytdn_20_columns.insert(52, 'BRANCH TYPE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column REFCAT3
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'REFCAT3']
ytdn_20_columns.insert(53, 'REFCAT3')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column BLP MONITORING
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'BLP MONITORING']
ytdn_20_columns.insert(54, 'BLP MONITORING')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column DS AREA
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'DS AREA']
ytdn_20_columns.insert(55, 'DS AREA')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column DS BAM CODE
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'DS BAM CODE']
ytdn_20_columns.insert(56, 'DS BAM CODE')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column DS BAM NAME
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'DS BAM NAME']
ytdn_20_columns.insert(57, 'DS BAM NAME')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column eKYC_RemoteSignature
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'eKYC_RemoteSignature']
ytdn_20_columns.insert(58, 'eKYC_RemoteSignature')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column eKYC_RemoteEPayment
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'eKYC_RemoteEPayment']
ytdn_20_columns.insert(59, 'eKYC_RemoteEPayment')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column MyData
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'MyData']
ytdn_20_columns.insert(60, 'MyData')
ytdn_20 = ytdn_20[ytdn_20_columns]

# Reordered column TCM MONITORING
ytdn_20_columns = [col for col in ytdn_20.columns if col != 'TCM MONITORING']
ytdn_20_columns.insert(61, 'TCM MONITORING')
ytdn_20 = ytdn_20[ytdn_20_columns]

ytdn_20.head()

# %% [markdown]
# ## Text Transformation

# %%
ytdn_21 = ytdn_20[['CA_CO', 'SUBMISSION DATE', 'SUBMISSION TIME', 'CA_POLNUM', 'CA_INAME',
       'SUSAMT1', 'APPROVAL DATE', 'ISSUE DATE', 'STATUS CODE', 'CA_BM',
       'CA_APREM', 'CA_SPREM', 'CA_QPREM', 'CA_MPREM', 'CB_COVAMT', 'CB_PLAN',
       'PN_PLAN_NAME', 'PHASE', 'OLD/NEW', 'CC_SERVAGT', 'EA_AGTNAME',
       'BRANCH CODE', 'BRANCH NAME', 'BANK DIVISION', 'BANK AREA',
       'BPI AIA DIVISION', 'BPI AIA AREA', 'BDM', 'BDM CODE', 'BAM CODE',
       'BAM', 'PRODUCT', 'ANP (SP @ 100%)', 'RP/SP', 'ANP in PESO',
       'REFERRER CODE', 'CLIENT SEGMENT', 'CLIENT SEGMENT 2', 'EXTRACT DATE',
       'BSE SEGMENT', 'Ok/Not Ok', 'CAMPAIGN CODE', 'TERRITORY', 'DPP Finder',
       'Pay Variant Finder', 'REG/Mat Recap', 'CC100 finder',
       'Pay Variant Finder 2', 'TCM finder', 'Pay Variant Finder 3', 'REFNAME',
       'REFCAT2', 'BRANCH TYPE', 'REFCAT3', 'BLP MONITORING', 'DS AREA',
       'DS BAM CODE', 'DS BAM NAME', 'eKYC_RemoteSignature',
       'eKYC_RemoteEPayment', 'MyData', 'TCM MONITORING']].copy()

ytdn_21['CA_INAME'] = ytdn_21['CA_INAME'].str.strip()
ytdn_21['CA_INAME'] = ytdn_21['CA_INAME'].apply(lambda x: ' '.join(x.split('/')[1:4]).replace('@', '')+'.')

ytdn_21['EA_AGTNAME'] = ytdn_21['EA_AGTNAME'].str.strip()
ytdn_21['EA_AGTNAME'] =ytdn_21['EA_AGTNAME'].apply(lambda x: ' '.join(x.split('/')[1:4]).replace('@', '')+'.')

ytdn_21.head()

# %% [markdown]
# ## COUNT and STATUS

# %%
ytdn_22 = ytdn_21.copy()

ytdn_22.insert(39, 'STATUS', 0)
ytdn_22.insert(39, 'COUNT', 0)

condition = [(ytdn_22['STATUS CODE'] == "00"), (ytdn_22['STATUS CODE'] == "10"), (ytdn_22['STATUS CODE'] == "0"), (ytdn_22['STATUS CODE'] == "13"), (ytdn_22['STATUS CODE'] == "14"), (ytdn_22['STATUS CODE'] == "15"), (ytdn_22['STATUS CODE'] == "16"), (ytdn_22['STATUS CODE'] == "43")]
choicelist = ["Pending", "Pending", "Pending", "Declined", "Incomplete", "Postponed", "Cancelled", "Surrender"]
ytdn_22['STATUS'] = np.select(condition, choicelist, "Approved")

ytdn_22['COUNT'] = 1

ytdn_22.head()

# %% [markdown]
# ## BPI-AIA Area and BPI-AIA On RM

# %%
ytdn_23 = ytdn_22.copy()

temp_df = db_bse_masterfile.drop_duplicates(subset=['BSE CODE']) # Remove duplicates so lookup merge only returns first match
db_bse_masterfile_tmp = temp_df[['BSE CODE','BPI-AIA DIVISION', 'BPI-AIA AREA']]

ytdn_23['CC_SERVAGT'] = ytdn_23['CC_SERVAGT'].str.strip()
df_merge = ytdn_23.merge(db_bse_masterfile_tmp, left_on=['CC_SERVAGT'], right_on=['BSE CODE'], how='left', suffixes=["", '_db_bse_masterfile'])
#df_merge_via_bsecode = df_merge.copy()

# Double Checking - Fill NaN Value during lookup using Mranch Code with Valuse Coming from BSE Code in BSE Table

df_merge['BPI AIA AREA'] = np.where(df_merge['REFCAT2'] == 'RM', df_merge['BPI-AIA AREA'], df_merge['BPI AIA AREA'])

df_merge['BPI AIA DIVISION'] = np.where(df_merge['REFCAT2'] == 'RM', df_merge['BPI-AIA DIVISION'], df_merge['BPI AIA DIVISION'])

ytdn_23 = df_merge.copy()

ytdn_23.head()

# %% [markdown]
# ## Reformat Time

# %%
ytdn_24 = ytdn_23.copy()

ytdn_24['SUBMISSION DATE'] = pd.to_datetime(ytdn_24['SUBMISSION DATE']).dt.strftime('%m/%d/%Y')
ytdn_24['APPROVAL DATE'] = pd.to_datetime(ytdn_24['APPROVAL DATE']).dt.strftime('%Y%m%d')
ytdn_24['ISSUE DATE'] = pd.to_datetime(ytdn_24['ISSUE DATE']).dt.strftime('%y%m%d')
ytdn_24['EXTRACT DATE'] = pd.to_datetime(ytdn_24['EXTRACT DATE']).dt.strftime('%m/%d/%Y')



# %% [markdown]
# ## Adding Bank Category

# %%
ytdn_25 = ytdn_24.copy()
ytdn_25["BANK CATEGORY"] = ''


## Doing a Bank Category

Selection = [ytdn_25['BRANCH CODE'].str.startswith('5'), ytdn_25['BPI AIA DIVISION'] == 'ABG',  ytdn_25['BPI AIA DIVISION'] =='CORPSOL']
Choices = ['RBANK', 'ABG', 'CORPSOL']

ytdn_25["BANK CATEGORY"] = np.select(Selection, Choices, 'BPI')


ytdn_25["BANK DIVISION"] = np.where(((ytdn_25["BRANCH CODE"] == '040000009') | (ytdn_25["BRANCH CODE"] == '710000001')) & (ytdn_25["BSE SEGMENT"] == 'ABG') , "ABG", ytdn_25["BANK DIVISION"])

ytdn_25["BANK AREA"] = np.where(((ytdn_25["BRANCH CODE"] == '040000009') | (ytdn_25["BRANCH CODE"] == '710000001')) & (ytdn_25["BSE SEGMENT"] == 'ABG') , "ABG", ytdn_25["BANK AREA"])



# %% [markdown]
# ## Special PB Condition

# %%
ytdn_25_1 = ytdn_25.copy()

ytdn_25_1['BANK DIVISION'] = np.where(((ytdn_25_1['BSE SEGMENT']=='PB') & (ytdn_25_1['BANK DIVISION']=='SG')), "PB", ytdn_25_1['BANK DIVISION'])

ytdn_25_1['BANK AREA'] = np.where(((ytdn_25_1['BSE SEGMENT']=='PB') & (ytdn_25_1['BANK AREA']=='SG')), "PB", ytdn_25_1['BANK AREA'])

ytdn_25_1['BPI AIA DIVISION'] = np.where(((ytdn_25_1['BSE SEGMENT']=='PB') & (ytdn_25_1['BPI AIA DIVISION']=='SG')), "PB", ytdn_25_1['BPI AIA DIVISION'])

ytdn_25_1['BPI AIA AREA'] = np.where(((ytdn_25_1['BSE SEGMENT']=='PB') & (ytdn_25_1['BPI AIA AREA']=='SG')), "PB", ytdn_25_1['BPI AIA AREA'])

# %% [markdown]
# ## Fill In Data Build Plus Pay

# %%
ytdn_26 = ytdn_25.copy()
ytdn_26['BLP PAY VARIANT'] = ''
ytdn_26['BLP PAY VARIANT'] = np.where(ytdn_26['PN_PLAN_NAME'].str.contains('Build Life Plus'), (ytdn_26['PN_PLAN_NAME'].str.split().apply(lambda x: " ".join(x[::-3]))), '-')
ytdn_26['BLP PAY VARIANT'] = ytdn_26['BLP PAY VARIANT'].str.replace('Build', '')
ytdn_26.head()

# ytdn_26.loc[ytdn_26['PN_PLAN_NAME'].str.contains('Build Life Plus')]
# ytdn_26.loc[ytdn_26['BANK CATEGORY'] == 'RBANK', ['BANK DIVISION', 'BANK AREA', 'BPI AIA DIVISION', 'BPI AIA AREA', 'TERRITORY']] = 'RBANK'
# ytdn_26

# %% [markdown]
# ## 50% ASPIRE PREM

# %%
ytdn_26_1 = ytdn_26.copy()

ytdn_26_1['ANP in PESO ORIGINAL'] = ytdn_26_1['ANP in PESO']
ytdn_26_1['ANP in PESO'] = np.where(ytdn_26_1['PRODUCT'] == 'ASPIRE PREMIER', (ytdn_26_1['ANP in PESO'] * 0.5), (ytdn_26_1['ANP in PESO']))

ytdn_26_1.head()

# %% [markdown]
# ## MTD Reference Data Transformation

# %% [markdown]
# ### Load the PrevMTD REF

# %%
datetime.now().strftime("%m%d%Y_%H%M%S")

# %%
now_year = datetime.now().strftime('%Y')                                

query = f'''
    SELECT * FROM [db_MTD_Ref_2024].[dbo].[tbl_MTD_Reference_Report_{now_year}]
'''
db_mtdref_old = sql_connection(query, 'db_MTD_Ref_2024')

db_mtdref_old = db_mtdref_old.loc[:,db_mtdref_old.columns!= 'DateUpload' ]
db_mtdref_old.columns

bacup_date = datetime.now().strftime("%m%d%Y_%H%M%S")

db_mtdref_old.to_excel(f"MTDRef_BackUp_{bacup_date}.xlsx", index=False)

# %%
# dtype_old = {
#     'CA_CO':str, 
#     'CA_POLNUM':str, 
#     'CA_INAME':str,
#     'SUSAMT1':str, 
#     'STATUS CODE':str, 
#     'CA_BM':str, 
#     'CB_PLAN':str,
#     'PN_PLAN_NAME':str, 
#     'PHASE':str, 
#     'OLD/NEW':str, 
#     'CC_SERVAGT':str, 
#     'EA_AGTNAME':str,
#     'BRANCH CODE':str, 
#     'BRANCH NAME':str, 
#     'BANK DIVISION':str, 
#     'BANK AREA':str,
#     'BPI AIA DIVISION':str, 
#     'BPI AIA AREA':str, 
#     'BDM':str, 
#     'BDM CODE':str, 
#     'BAM CODE':str,
#     'BAM':str, 
#     'PRODUCT':str, 
#     'RP/SP':str, 
#     'REFERRER CODE':str, 
#     'CLIENT SEGMENT':str, 
#     'CLIENT SEGMENT 2':str, 
#     'STATUS':str, 
#     'BSE SEGMENT':str, 
#     'Ok/Not Ok':str, 
#     'CAMPAIGN CODE':str, 
#     'TERRITORY':str,
#     'DPP Finder':str, 
#     'Pay Variant Finder':str,
#     'REG/Mat Recap':str, 
#     'CC100 finder':str,
#     'Pay Variant Finder 2':str,
#     'TCM finder':str, 
#     'Pay Variant Finder 3':str, 
#     'REFNAME':str,
#     'REFCAT2':str,
#     'BRANCH TYPE':str, 
#     'REFCAT3':str, 
#     'BLP MONITORING':str, 
#     'DS AREA':str,
#     'DS BAM CODE':str, 
#     'DS BAM NAME':str, 
#     'eKYC_RemoteSignature':str,
#     'eKYC_RemoteEPayment':str, 
#     'MyData':str, 
#     'TCM MONITORING':str
# }
# db_mtdref_old = pd.read_excel(r'C:\Users\i024605\Documents\MTD_20240220 Reference with T1.xlsx', sheet_name='YTD Reference', dtype=dtype_old)
# db_mtdref_old.shape
# db_mtdref_old.columns

# %% [markdown]
# ### Perform Lookup and Filtering to Prev MTDRef 

# %%
db_mtdref_old_1 = db_mtdref_old.copy()

##Applying BPLP Even in the Old Records
db_mtdref_old_1['BLP PAY VARIANT'] = ''
db_mtdref_old_1['BLP PAY VARIANT'] = np.where(db_mtdref_old_1['PN_PLAN_NAME'].str.contains('Build Life Plus'), (db_mtdref_old_1['PN_PLAN_NAME'].str.split().apply(lambda x: " ".join(x[::-3]))), '-')
db_mtdref_old_1['BLP PAY VARIANT'] = db_mtdref_old_1['BLP PAY VARIANT'].str.replace('Build', '')
db_mtdref_old_1

## Doing a Bank Category

# Selection = [db_mtdref_old_1['BRANCH CODE'].str.startswith('5'), db_mtdref_old_1['BPI AIA DIVISION'] == 'ABG',  db_mtdref_old_1['BPI AIA DIVISION'] =='CORPSOL']
# Choices = ['RBANK', 'ABG', 'CORPSOL']

# db_mtdref_old_1["BANK CATEGORY"] = np.select(Selection, Choices, 'BPI').astype(bool)
# db_mtdref_old_1


db_mtdref_old_1["BANK CATEGORY"] = 'BPI'
db_mtdref_old_1["BANK CATEGORY"] = np.where(db_mtdref_old_1['BRANCH CODE'].str.startswith('5'), "RBANK", db_mtdref_old_1["BANK CATEGORY"])
db_mtdref_old_1["BANK CATEGORY"] = np.where(db_mtdref_old_1['BPI AIA DIVISION'] == 'ABG', "ABG", db_mtdref_old_1["BANK CATEGORY"])
db_mtdref_old_1["BANK CATEGORY"] = np.where(db_mtdref_old_1['BPI AIA DIVISION'] == 'CORPSOL', "CORPSOL", db_mtdref_old_1["BANK CATEGORY"])
#db_mtdref_old_1["BANK CATEGORY"] = np.where(((db_mtdref_old_1["BRANCH CODE"] == '040000009') | (db_mtdref_old_1["BRANCH CODE"] == '710000001')), "ABG", db_mtdref_old_1["BANK CATEGORY"])
db_mtdref_old_1["BANK DIVISION"] = np.where(((db_mtdref_old_1["BRANCH CODE"] == '040000009') | (db_mtdref_old_1["BRANCH CODE"] == '710000001')) & (db_mtdref_old_1["BSE SEGMENT"] == 'ABG') , "ABG", db_mtdref_old_1["BANK DIVISION"])
db_mtdref_old_1["BANK AREA"] = np.where(((db_mtdref_old_1["BRANCH CODE"] == '040000009') | (db_mtdref_old_1["BRANCH CODE"] == '710000001')) & (db_mtdref_old_1["BSE SEGMENT"] == 'ABG') , "ABG", db_mtdref_old_1["BANK AREA"])

# db_mtdref_old_1 = db_mtdref_old_1.copy()
# db_mtdref_old_1.loc[db_mtdref_old_1['BANK CATEGORY'] == 'RBANK', ['BANK DIVISION', 'BANK AREA', 'BPI AIA DIVISION', 'BPI AIA AREA', 'TERRITORY']] = 'RBANK'
# db_mtdref_old_1

ytdn_27 = ytdn_26_1[['CA_CO', 'SUBMISSION DATE', 'SUBMISSION TIME', 'CA_POLNUM', 'CA_INAME',
       'SUSAMT1', 'APPROVAL DATE', 'ISSUE DATE', 'STATUS CODE', 'CA_BM',
       'CA_APREM', 'CA_SPREM', 'CA_QPREM', 'CA_MPREM', 'CB_COVAMT', 'CB_PLAN',
       'PN_PLAN_NAME', 'PHASE', 'OLD/NEW', 'CC_SERVAGT', 'EA_AGTNAME',
       'BRANCH CODE', 'BRANCH NAME', 'BANK DIVISION', 'BANK AREA',
       'BPI AIA DIVISION', 'BPI AIA AREA', 'BDM CODE', 'BDM', 'BAM CODE',
       'BAM', 'PRODUCT', 'ANP (SP @ 100%)', 'RP/SP', 'ANP in PESO',
       'REFERRER CODE', 'CLIENT SEGMENT', 'CLIENT SEGMENT 2', 'EXTRACT DATE',
       'COUNT', 'STATUS', 'BSE SEGMENT', 'Ok/Not Ok', 'CAMPAIGN CODE', 'TERRITORY',
       'DPP Finder', 'Pay Variant Finder', 'REG/Mat Recap', 'CC100 finder',
       'Pay Variant Finder 2', 'TCM finder', 'Pay Variant Finder 3', 'REFNAME',
       'REFCAT2', 'BRANCH TYPE', 'REFCAT3', 'BLP MONITORING', 'DS AREA',
       'DS BAM CODE', 'DS BAM NAME', 'eKYC_RemoteSignature',
       'eKYC_RemoteEPayment', 'MyData', 'TCM MONITORING', 'BANK CATEGORY', 'BLP PAY VARIANT', 'ANP in PESO ORIGINAL']].copy()

## Formatting Extract Date for Prev Extract Date
db_mtdref_old_1['EXTRACT DATE'] = pd.to_datetime(db_mtdref_old_1['EXTRACT DATE']).dt.strftime('%m/%d/%Y')

## Remove any AGTNAME that have IPOS TEAM
ytdn_27 = ytdn_27[~ytdn_27['EA_AGTNAME'].str.contains('IPOS TEAM')]
ytdn_27 = ytdn_27[~ytdn_27['CA_POLNUM'].str.contains('7012744487')]


print(ytdn_27.shape)
ytdn_27.head()

# %%
# db_mtdref_old.loc[db_mtdref_old['EXTRACT DATE'] == '01/31/2024']
#ytdn_27.loc[ytdn_27['BANK CATEGORY']=='ABG']
#ytdn_27.loc[ytdn_27['CA_POLNUM'] == '7012756752']


# %% [markdown]
# ### Normal Days
# #### With Time Filter

# %%
def normal_days_calling():
    recorded_policy = np.array(db_mtdref_old_1['CA_POLNUM'].drop_duplicates())

    db_mtdref_not_currentmonth = db_mtdref_old_1.loc[pd.to_datetime(db_mtdref_old_1['EXTRACT DATE']).dt.strftime("%m%Y") != datetime.now().strftime("%m%Y")]
    db_mtdref_currentmonth = db_mtdref_old_1.loc[pd.to_datetime(db_mtdref_old_1['EXTRACT DATE']).dt.strftime("%m%Y") == datetime.now().strftime("%m%Y")]

    print("db_mtdref_not_currentmonth: ", db_mtdref_not_currentmonth.shape)
    print("db_mtdref_currentmonth: ", db_mtdref_currentmonth.shape)

    #### Getting the new Entry
    ytdn_27_temp = ytdn_27.copy()

    temp_df = db_mtdref_old_1.drop_duplicates(subset=['CA_POLNUM']) # Remove duplicates so lookup merge only returns first match
    temp_df = temp_df[['CA_POLNUM', 'EXTRACT DATE']]
    ytdn_27_temp['CA_POLNUM'] = ytdn_27_temp['CA_POLNUM'].str.strip()
    ytdn_27_temp = ytdn_27_temp.merge(temp_df, left_on=['CA_POLNUM'], right_on=['CA_POLNUM'], how='left', suffixes=['', '_db_mtdref_old_1'])


    # ytdn_new_enrty = ytdn_27_temp.loc[ytdn_27_temp['EXTRACT DATE_db_mtdref_old_1'].isna()]
    ytdn_new_enrty = ytdn_27_temp[~ytdn_27_temp['CA_POLNUM'].isin(recorded_policy)]
    ytdn_new_enrty = ytdn_new_enrty.drop('EXTRACT DATE_db_mtdref_old_1', axis=1)

    ytdn_new_enrty

    #### Removing the "Not Current Months" in the YTDN.
    ### This was performed to remove not current months as we only need to update those who are in the current month
    ytdn_27_temp = ytdn_27.copy()
    temp_df = db_mtdref_not_currentmonth.drop_duplicates(subset=['CA_POLNUM']).copy() # Remove duplicates so lookup merge only returns first match
    temp_df = temp_df[['CA_POLNUM', 'EXTRACT DATE']]
    ytdn_27_temp['CA_POLNUM'] = ytdn_27_temp['CA_POLNUM'].str.strip()
    ytdn_27_temp = ytdn_27_temp.merge(temp_df, left_on=['CA_POLNUM'], right_on=['CA_POLNUM'], how='left', suffixes=['', '_db_mtdref_not_currentmonth'])


    ytdn_remove_not_current_month_enrty = ytdn_27_temp.loc[ytdn_27_temp['EXTRACT DATE_db_mtdref_not_currentmonth'].isna()]
    ytdn_remove_not_current_month_enrty = ytdn_remove_not_current_month_enrty.drop('EXTRACT DATE_db_mtdref_not_currentmonth', axis=1)

    ytdn_remove_not_current_month_enrty

    #### Removing the Old Current Month
    ytdn_27_temp = ytdn_remove_not_current_month_enrty.copy()

    temp_df = db_mtdref_currentmonth.drop_duplicates(subset=['CA_POLNUM']).copy() # Remove duplicates so lookup merge only returns first match
    temp_df = temp_df[['CA_POLNUM', 'EXTRACT DATE']]
    ytdn_27_temp['CA_POLNUM'] = ytdn_27_temp['CA_POLNUM'].str.strip()
    ytdn_27_temp = ytdn_27_temp.merge(temp_df, left_on=['CA_POLNUM'], right_on=['CA_POLNUM'], how='left', suffixes=['', '_db_mtdref_currentmonth'])


    ytdn_remove_existing_current_month_enrty = ytdn_27_temp.loc[ytdn_27_temp['EXTRACT DATE_db_mtdref_currentmonth'].isna()]
    ytdn_remove_existing_current_month_enrty = ytdn_remove_existing_current_month_enrty.drop('EXTRACT DATE_db_mtdref_currentmonth', axis=1)

    print("ytdn_new_enrty: ", ytdn_new_enrty.shape)
    print('ytdn_remove_existing_current_month_enrty: ', ytdn_remove_existing_current_month_enrty.shape)

    ## Comparison Checker
    if (ytdn_remove_existing_current_month_enrty.shape[0] == ytdn_new_enrty.shape[0]) & (ytdn_remove_existing_current_month_enrty.shape[1] == ytdn_new_enrty.shape[1]):
        print("Number of Columns Matches Proceed")
        ytdn_daily_1 = ytdn_remove_existing_current_month_enrty
    else:
        print("Please Check Data")

    ##Take Only Records Before 4:05 PM 
    #ytdn_daily_1 = ytdn_daily_1.loc[ytdn_daily_1['SUBMISSION TIME'] < '16:05:00']

    ## Transfer the extract date only from old to current
    db_mtdref_currentmonth_update = ytdn_27_temp.loc[ytdn_27_temp['EXTRACT DATE_db_mtdref_currentmonth'].notna()]
    db_mtdref_currentmonth_update['EXTRACT DATE'] = db_mtdref_currentmonth_update['EXTRACT DATE_db_mtdref_currentmonth']
    db_mtdref_currentmonth_update = db_mtdref_currentmonth_update.drop('EXTRACT DATE_db_mtdref_currentmonth', axis=1)

    db_mtdref_currentmonth_update

    ## By default the number of columns is 64 due to INAME being removed once roll-out to production

    print("Finish Function Run")

    return db_mtdref_not_currentmonth, db_mtdref_currentmonth_update, ytdn_daily_1

# db_mtdref_not_currentmonth, db_mtdref_currentmonth_update, ytdn_daily_1 = normal_days_calling()

# %% [markdown]
# ### 1st Day of the Month Condition
# #### With Time Filter
# 

# %%
 
def first_day_month_calling():

    recorded_policy = np.array(db_mtdref_old_1['CA_POLNUM'].drop_duplicates())

    today = date.today()

    yesteday = today - timedelta(days=1)
    yesteday.strftime("%m%Y")

    date_2MonthsOld = today + relativedelta(months=-1)
    date_2MonthsOld.strftime("%m%Y")

    # print((date.today() + relativedelta(months=-2)).strftime("%m%Y"))
    # print(datetime.today().replace(day=1).strftime('%m%d%y'))
    # print(datetime.now().strftime('%m%d%y'))

    # if datetime.now().strftime('%m%d%y') == datetime.today().replace(day=1).strftime('%m%d%y'):
    #     print("Get the Get the last month")
    # else:
    #     print("Continue")


    db_mtdref_morethan2months_old= db_mtdref_old_1.loc[pd.to_datetime(db_mtdref_old_1['EXTRACT DATE']).dt.strftime("%m%Y") < (date.today() + relativedelta(months=-1)).strftime("%m%Y")]
    db_mtdref_1months_old = db_mtdref_old_1.loc[pd.to_datetime(db_mtdref_old_1['EXTRACT DATE']).dt.strftime("%m%Y") == (date.today() + relativedelta(months=-1)).strftime("%m%Y")]

    print("db_mtdref_not_currentmonth: ", db_mtdref_morethan2months_old.shape)
    print("db_mtdref_1months_old: ", db_mtdref_1months_old.shape)

    ## 1st Stemp is to Remove File Older then 1 Month
    ytdn_27_temp = ytdn_27.copy()
    temp_df = db_mtdref_morethan2months_old.drop_duplicates(subset=['CA_POLNUM']) # Remove duplicates so lookup merge only returns first match
    temp_df = temp_df[['CA_POLNUM', 'EXTRACT DATE']]
    ytdn_27_temp['CA_POLNUM'] = ytdn_27_temp['CA_POLNUM'].str.strip()
    ytdn_27_temp = ytdn_27_temp.merge(temp_df, left_on=['CA_POLNUM'], right_on=['CA_POLNUM'], how='left', suffixes=['', '_db_mtdref_old_1'])

    ytdn_remove_older1month = ytdn_27_temp.loc[ytdn_27_temp['EXTRACT DATE_db_mtdref_old_1'].isna()]
    ytdn_remove_older1month = ytdn_27_temp.drop('EXTRACT DATE_db_mtdref_old_1', axis=1)
    print("ytdn_remove_older1month: ", ytdn_remove_older1month.shape)


    # 2nd Step is to modify Data 1 Month Old
    ytdn_27_temp = ytdn_remove_older1month.copy()
    temp_df = db_mtdref_1months_old.drop_duplicates(subset=['CA_POLNUM']) # Remove duplicates so lookup merge only returns first match
    temp_df = temp_df[['CA_POLNUM', 'EXTRACT DATE']]
    ytdn_27_temp['CA_POLNUM'] = ytdn_27_temp['CA_POLNUM'].str.strip()
    ytdn_27_temp = ytdn_27_temp.merge(temp_df, left_on=['CA_POLNUM'], right_on=['CA_POLNUM'], how='left', suffixes=['', '_db_mtdref_1months_old'])

    ytdn_1monthold_mod = ytdn_27_temp.loc[ytdn_27_temp['EXTRACT DATE_db_mtdref_1months_old'].notna()] ## Getting The 1 Month Old Data Modified Only

    ytdn_1monthold_mod['EXTRACT DATE'] = ytdn_1monthold_mod['EXTRACT DATE_db_mtdref_1months_old']  ## Updating the Whole 1 Month Old Data Modified Besides Extract Date
    ytdn_1monthold_mod = ytdn_1monthold_mod.drop('EXTRACT DATE_db_mtdref_1months_old', axis=1) ## Drop EXTRACT DATE_db_mtdref_1months_old
    print("ytdn_1monthold_mod: ", ytdn_1monthold_mod.shape)

    ## Get the only Updated Data 
    ytdn_current_data = ytdn_27_temp.loc[ytdn_27_temp['EXTRACT DATE_db_mtdref_1months_old'].isna()] 
    print("ytdn_current_data: ", ytdn_current_data.shape)

    ytdn_current_data_checker = ytdn_27_temp[~ytdn_27_temp['CA_POLNUM'].isin(recorded_policy)]
        ## Comparison Checker
    if (ytdn_current_data.shape[0] == ytdn_current_data_checker.shape[0]) & (ytdn_current_data.shape[1] == ytdn_current_data_checker.shape[1]):
        print("Number of Columns Matches Proceed")
        ytdn_current_data_checker = ytdn_current_data
    else:
        print("Please Check Data")
        #raise Exception("Error")

    ## Performning Comparison of Submission Date VS Extract Date
    yesteday_date = (today - timedelta(days=1)).strftime("%m/%d/%Y")
    yesteday_date

    today_date = today.strftime("%m/%d/%Y")
    today_date

    ytdn_current_data['EXTRACT DATE'] = np.where((ytdn_current_data['EXTRACT DATE'] > ytdn_current_data['SUBMISSION DATE']), yesteday_date, today_date )
    
    ##Take Only Records Before 4:05 PM 
    # ytdn_current_data = ytdn_current_data.loc[ytdn_current_data['SUBMISSION TIME'] < '16:05:00']

    print("Finish Function Run")
    return db_mtdref_morethan2months_old, ytdn_1monthold_mod, ytdn_current_data

# %% [markdown]
# ### Selection Performing Concatination of MTD Ref

# %%
### Combining the 3 DataFrame the 1. Not in Current Month, 2. Current Month Updated, 3. New Entry
if datetime.now().strftime('%m%d%y') != datetime.today().replace(day=1).strftime('%m%d%y'):

       print("Normal Days")

       db_mtdref_not_currentmonth, db_mtdref_currentmonth_update, ytdn_daily_1 = normal_days_calling()
       

       db_mtdref_not_currentmonth['CA_INAME'] = ''

       try:
              db_mtdref_not_currentmonth = db_mtdref_not_currentmonth[['CA_CO', 'SUBMISSION DATE', 'SUBMISSION TIME', 'CA_POLNUM', 'CA_INAME', 'SUSAMT1',
                     'APPROVAL DATE', 'ISSUE DATE', 'STATUS CODE', 'CA_BMODE', 'CA_APREM',
                     'CA_SPREM', 'CA_QPREM', 'CA_MPREM', 'CB_COVAMT', 'CB_PLAN',
                     'PN_PLAN_NAME', 'PHASE', 'OLD/NEW', 'CC_SERVAGT', 'EA_AGTNAME',
                     'BRANCH CODE', 'BRANCH NAME', 'BANK DIVISION', 'BANK AREA',
                     'BPI AIA DIVISION', 'BPI AIA AREA', 'BDM CODE', 'BDM', 'BAM CODE',
                     'BAM', 'PRODUCT', 'ANP (SP @ 100%)', 'RP/SP', 'ANP in Peso',
                     'REFERRER CODE', 'CLIENT SEGMENT', 'CLIENT SEGMENT 2', 'EXTRACT DATE',
                     'COUNT', 'STATUS', 'BSE SEGMENT', 'Ok/Not Ok', 'CAMPAIGN CODE',
                     'TERRITORY', 'DPP finder', 'Pay Variant Finder', 'REG/Mat Recap',
                     'CC100 finder', 'Pay Variant Finder.1', 'TCM finder',
                     'Pay Variant Finder.2', 'REFNAME', 'REFCAT2', 'BRANCH TYPE', 'REFCAT3',
                     'BLP MONITORING', 'DS AREA', 'DSBAM CODE', 'DS BAM NAME',
                     'eKYC_RemoteSignature', 'eKYC_RemoteEPayment', 'MyData',
                     'TCM MONITORING', 'BANK CATEGORY', 'BLP PAY VARIANT', 'ANP in PESO ORIGINAL']]
       except:
                   
              db_mtdref_not_currentmonth = db_mtdref_not_currentmonth[['CA_CO', 'SUBMISSION DATE', 'SUBMISSION TIME', 'CA_POLNUM', 'CA_INAME', 'SUSAMT1',
                     'APPROVAL DATE', 'ISSUE DATE', 'STATUS CODE', 'CA_BM', 'CA_APREM',
                     'CA_SPREM', 'CA_QPREM', 'CA_MPREM', 'CB_COVAMT', 'CB_PLAN',
                     'PN_PLAN_NAME', 'PHASE', 'OLD/NEW', 'CC_SERVAGT', 'EA_AGTNAME',
                     'BRANCH CODE', 'BRANCH NAME', 'BANK DIVISION', 'BANK AREA',
                     'BPI AIA DIVISION', 'BPI AIA AREA', 'BDM CODE', 'BDM', 'BAM CODE',
                     'BAM', 'PRODUCT', 'ANP (SP @ 100%)', 'RP/SP', 'ANP in PESO',
                     'REFERRER CODE', 'CLIENT SEGMENT', 'CLIENT SEGMENT 2', 'EXTRACT DATE',
                     'COUNT', 'STATUS', 'BSE SEGMENT', 'Ok/Not Ok', 'CAMPAIGN CODE',
                     'TERRITORY', 'DPP Finder', 'Pay Variant Finder', 'REG/Mat Recap',
                     'CC100 finder', 'Pay Variant Finder 2', 'TCM finder',
                     'Pay Variant Finder 3', 'REFNAME', 'REFCAT2', 'BRANCH TYPE', 'REFCAT3',
                     'BLP MONITORING', 'DS AREA', 'DS BAM CODE', 'DS BAM NAME',
                     'eKYC_RemoteSignature', 'eKYC_RemoteEPayment', 'MyData',
                     'TCM MONITORING', 'BANK CATEGORY', 'BLP PAY VARIANT', 'ANP in PESO ORIGINAL']]

       #Renamed Columns from the Old to match the new and perform concat
       db_mtdref_not_currentmonth.rename(columns = {'CA_BMODE':'CA_BM', 'Pay Variant Finder.1':'Pay Variant Finder 2', 'Pay Variant Finder.2': 'Pay Variant Finder 3',
                                                 'DSBAM CODE' : 'DS BAM CODE', 'ANP in Peso':'ANP in PESO', 'DPP finder': 'DPP Finder'}, inplace = True) 


       db_mtdref_currentmonth_update
       
       ytdn_daily_1

       #Temporary Fix
       # ytdn_daily_1['EXTRACT DATE'] = ytdn_daily_1['SUBMISSION DATE']

       ## Redeclaration Only So I wont get confused



       ytdn_28 = pd.concat([db_mtdref_not_currentmonth, db_mtdref_currentmonth_update, ytdn_daily_1], axis=0)

       Monlthy_Included_Policy = pd.concat([db_mtdref_currentmonth_update, ytdn_daily_1], axis=0)
       Monlthy_Included_Policy.drop_duplicates(subset=['CA_POLNUM'], keep='first', inplace=True)

       #print("Normal Days")

elif datetime.now().strftime('%m%d%y') == datetime.today().replace(day=1).strftime('%m%d%y'):
       print("Special Days")
       db_mtdref_morethan2months_old, ytdn_1monthold_mod, ytdn_current_data = first_day_month_calling()
       db_mtdref_not_currentmonth, db_mtdref_currentmonth_update, ytdn_daily_1 = first_day_month_calling()

       db_mtdref_morethan2months_old['CA_INAME'] = ''
       
       ## Using Excel
       try:
              db_mtdref_morethan2months_old = db_mtdref_morethan2months_old[['CA_CO', 'SUBMISSION DATE', 'SUBMISSION TIME', 'CA_POLNUM', 'CA_INAME', 'SUSAMT1',
                     'APPROVAL DATE', 'ISSUE DATE', 'STATUS CODE', 'CA_BMODE', 'CA_APREM',
                     'CA_SPREM', 'CA_QPREM', 'CA_MPREM', 'CB_COVAMT', 'CB_PLAN',
                     'PN_PLAN_NAME', 'PHASE', 'OLD/NEW', 'CC_SERVAGT', 'EA_AGTNAME',
                     'BRANCH CODE', 'BRANCH NAME', 'BANK DIVISION', 'BANK AREA',
                     'BPI AIA DIVISION', 'BPI AIA AREA', 'BDM CODE', 'BDM', 'BAM CODE',
                     'BAM', 'PRODUCT', 'ANP (SP @ 100%)', 'RP/SP', 'ANP in Peso',
                     'REFERRER CODE', 'CLIENT SEGMENT', 'CLIENT SEGMENT 2', 'EXTRACT DATE',
                     'COUNT', 'STATUS', 'BSE SEGMENT', 'Ok/Not Ok', 'CAMPAIGN CODE',
                     'TERRITORY', 'DPP finder', 'Pay Variant Finder', 'REG/Mat Recap',
                     'CC100 finder', 'Pay Variant Finder.1', 'TCM finder',
                     'Pay Variant Finder.2', 'REFNAME', 'REFCAT2', 'BRANCH TYPE', 'REFCAT3',
                     'BLP MONITORING', 'DS AREA', 'DSBAM CODE', 'DS BAM NAME',
                     'eKYC_RemoteSignature', 'eKYC_RemoteEPayment', 'MyData',
                     'TCM MONITORING', 'BANK CATEGORY', 'BLP PAY VARIANT', 'ANP in PESO ORIGINAL']]

       except:
              db_mtdref_morethan2months_old = db_mtdref_morethan2months_old[['CA_CO', 'SUBMISSION DATE', 'SUBMISSION TIME', 'CA_POLNUM', 'CA_INAME', 'SUSAMT1',
                     'APPROVAL DATE', 'ISSUE DATE', 'STATUS CODE', 'CA_BM', 'CA_APREM',
                     'CA_SPREM', 'CA_QPREM', 'CA_MPREM', 'CB_COVAMT', 'CB_PLAN',
                     'PN_PLAN_NAME', 'PHASE', 'OLD/NEW', 'CC_SERVAGT', 'EA_AGTNAME',
                     'BRANCH CODE', 'BRANCH NAME', 'BANK DIVISION', 'BANK AREA',
                     'BPI AIA DIVISION', 'BPI AIA AREA', 'BDM CODE', 'BDM', 'BAM CODE',
                     'BAM', 'PRODUCT', 'ANP (SP @ 100%)', 'RP/SP', 'ANP in PESO',
                     'REFERRER CODE', 'CLIENT SEGMENT', 'CLIENT SEGMENT 2', 'EXTRACT DATE',
                     'COUNT', 'STATUS', 'BSE SEGMENT', 'Ok/Not Ok', 'CAMPAIGN CODE',
                     'TERRITORY', 'DPP Finder', 'Pay Variant Finder', 'REG/Mat Recap',
                     'CC100 finder', 'Pay Variant Finder 2', 'TCM finder',
                     'Pay Variant Finder 3', 'REFNAME', 'REFCAT2', 'BRANCH TYPE', 'REFCAT3',
                     'BLP MONITORING', 'DS AREA', 'DS BAM CODE', 'DS BAM NAME',
                     'eKYC_RemoteSignature', 'eKYC_RemoteEPayment', 'MyData',
                     'TCM MONITORING', 'BANK CATEGORY', 'BLP PAY VARIANT', 'ANP in PESO ORIGINAL']]

       #Renamed Columns from the Old to match the new and perform concat
       db_mtdref_morethan2months_old.rename(columns = {'CA_BMODE':'CA_BM', 'Pay Variant Finder.1':'Pay Variant Finder 2', 'Pay Variant Finder.2': 'Pay Variant Finder 3',
                                          'DSBAM CODE' : 'DS BAM CODE', 'ANP in Peso':'ANP in PESO', 'DPP finder': 'DPP Finder'}, inplace = True) 

       

       ytdn_1monthold_mod

       ytdn_current_data

       ## Redeclaration Only So I wont get confused
       

       ytdn_28 = pd.concat([db_mtdref_morethan2months_old, ytdn_1monthold_mod, ytdn_current_data], axis=0)

       Monlthy_Included_Policy = pd.concat([ytdn_1monthold_mod, ytdn_current_data], axis=0)
       Monlthy_Included_Policy.drop_duplicates(subset=['CA_POLNUM'], keep='first', inplace=True)
       # print("Special Days")

else:
       print("Check the code")
       

# %%
print(db_mtdref_not_currentmonth.shape)
print(db_mtdref_currentmonth_update.shape)
print(ytdn_daily_1.shape)


# %%
# np.array(ytdn_daily_1['CA_POLNUM'])

# %%
ytdn_28.loc[ytdn_28['PRODUCT'].isna()]

# %% [markdown]
# ## Removing Policy Number Possible Duplicates

# %%
ytdn_29 = ytdn_28.copy()
ytdn_29.drop_duplicates(subset=['CA_POLNUM'], keep='first', inplace=True)
ytdn_29

# %% [markdown]
# ## MTDRef_Database

# %%
MTD_Reference_Database = ytdn_29[['CA_CO', 'SUBMISSION DATE', 'SUBMISSION TIME', 'CA_POLNUM', 'CA_INAME',
       'SUSAMT1', 'APPROVAL DATE', 'ISSUE DATE', 'STATUS CODE', 'CA_BM',
       'CA_APREM', 'CA_SPREM', 'CA_QPREM', 'CA_MPREM', 'CB_COVAMT', 'CB_PLAN',
       'PN_PLAN_NAME', 'PHASE', 'OLD/NEW', 'CC_SERVAGT', 'EA_AGTNAME',
       'BRANCH CODE', 'BRANCH NAME', 'BANK DIVISION', 'BANK AREA',
       'BPI AIA DIVISION', 'BPI AIA AREA', 'BDM CODE', 'BDM', 'BAM CODE',
       'BAM', 'PRODUCT', 'ANP (SP @ 100%)', 'RP/SP', 'ANP in PESO',
       'REFERRER CODE', 'CLIENT SEGMENT', 'CLIENT SEGMENT 2', 'EXTRACT DATE',
       'COUNT', 'STATUS', 'BSE SEGMENT', 'Ok/Not Ok', 'CAMPAIGN CODE', 'TERRITORY',
       'DPP Finder', 'Pay Variant Finder', 'REG/Mat Recap', 'CC100 finder',
       'Pay Variant Finder 2', 'TCM finder', 'Pay Variant Finder 3', 'REFNAME',
       'REFCAT2', 'BRANCH TYPE', 'REFCAT3', 'BLP MONITORING', 'DS AREA',
       'DS BAM CODE', 'DS BAM NAME', 'eKYC_RemoteSignature',
       'eKYC_RemoteEPayment', 'MyData', 'TCM MONITORING', "BANK CATEGORY", "BLP PAY VARIANT",'ANP in PESO ORIGINAL']].copy()

MTD_Reference_Database['SUBMISSION DATE'] = (pd.to_datetime(MTD_Reference_Database['SUBMISSION DATE'])).dt.strftime('%m/%d/%Y')

MTD_Obj_Col = MTD_Reference_Database.select_dtypes(include=['object']).columns
MTD_Reference_Database[MTD_Obj_Col] = MTD_Reference_Database[MTD_Obj_Col].applymap(lambda x: x.strip() if isinstance(x, str) else x)

MTD_Reference_Database

# %% [markdown]
# ## For Sharing

# %%
MTD_Reference_Public = MTD_Reference_Database.loc[:,MTD_Reference_Database.columns!='CA_INAME']
MTD_Reference_Public

# %% [markdown]
# ## Converting DateTime64 For Excel Date Format

# %%
MTD_Reference_Public['SUBMISSION DATE'] = MTD_Reference_Public['SUBMISSION DATE'].astype('datetime64[ns]')
MTD_Reference_Public['EXTRACT DATE'] = MTD_Reference_Public['EXTRACT DATE'].astype('datetime64[ns]')

ytdn_daily_1['SUBMISSION DATE'] = ytdn_daily_1['SUBMISSION DATE'].astype('datetime64[ns]')
ytdn_daily_1['EXTRACT DATE'] = ytdn_daily_1['EXTRACT DATE'].astype('datetime64[ns]')

# %%
# MTD_Reference_Public.fillna('')
# ytdn_daily_1.fillna('')

# Pol_tobe_Included = Monlthy_Included_Policy['CA_POLNUM'].copy()
# Pol_tobe_Included = np.array(Pol_tobe_Included)
# Pol_tobe_Included

# MTD_Reference_Report = MTD_Reference_Public.reset_index(drop=True).copy()
# MTD_Reference_Report = MTD_Reference_Report[MTD_Reference_Report['CA_POLNUM'].isin(Pol_tobe_Included)]

# MTD_Reference_Report.to_excel('Reference with T1 Automate 021324.xlsx', index=False)

# %%
MTD_Reference_Public.loc[((MTD_Reference_Public["BRANCH CODE"] == '040000009') | (MTD_Reference_Public["BRANCH CODE"] == '710000001')) & (MTD_Reference_Public["BSE SEGMENT"] == 'ABG')].shape
# MTD_Reference_Public.loc[MTD_Reference_Public['EXTRACT DATE']== '2024-06-30', 'EXTRACT DATE'] = '2024-06-28'

# %%
print(MTD_Reference_Public['ANP in PESO'].sum())
print(MTD_Reference_Public['ANP in PESO ORIGINAL'].sum())

# %% [markdown]
# # Pamilya Protect

# %%
## Import System Library
import os
import glob
import shutil
import warnings
import time
from time import sleep

## Import Data Manipulation Library
import pandas as pd
import numpy as np
from sqlalchemy import create_engine
import pyodbc 
import warnings
from datetime import datetime, timedelta

## Import WebScrap Library
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

## Set the Maximum Column Display to Max and Ignore any warnings
pd.set_option('display.max_columns', None)
warnings.filterwarnings('ignore')

# %%
# directory = r'C:\Users\I024605\Downloads'
directory = os.path.join(os.path.expanduser("~"), "Downloads")

list_of_files = glob.glob(directory+'\\Rapid Raw Data_Latest*.xlsb')

try:
    if os.path.isfile(list_of_files[0]):
        os.remove(list_of_files[0])
        print("File Removed")
except:
    print("No File Found")
    pass

sleep(2)

# %%
directory = os.path.join(os.path.expanduser("~"), "Downloads\PProtect_Data")

list_of_files = glob.glob(directory+'\\Rapid Raw Data_Latest*.xlsb')

try:
    if os.path.isfile(list_of_files[0]):
        os.remove(list_of_files[0])
        print("File Removed")
except:
    print("No File Found")
    pass

sleep(2)

# %% [markdown]
# ## Make a WebScrap Function

# %%
def WebScrap():
    # chrome_options = Options()
    # chrome_options.add_argument("--headless=new")

    chrome_options = Options()
    chrome_options.add_argument("--headless")

    import time

    #driver = webdriver.Chrome(options=chrome_options)

    driver = webdriver.Chrome()

    driver.get("https://aiacom.sharepoint.com/:x:/r/sites/ProjectFirestormPH/Shared%20Documents/Reports/Raw%20Data/Rapid%20Raw%20Data_Latest.xlsb")

    try:
        element_present = EC.presence_of_element_located((By.ID, 'i0116'))
        WebDriverWait(driver, 10).until(element_present)

    except TimeoutException:
        print ("Timed out waiting for page to load")

    driver.find_element(By.ID, 'i0116').send_keys('AshnerGerald.Novilla@aia.com')

    try:
        element_present = EC.presence_of_element_located((By.ID, 'i0116'))
        WebDriverWait(driver, 10).until(element_present)

    except TimeoutException:
        print ("Timed out waiting for page to load")


    driver.find_element(By.ID, 'i0116').send_keys(Keys.ENTER)

    time.sleep(10)

    os.chdir(os.path.join(os.path.expanduser("~"), "Downloads"))

    dir_list = pd.DataFrame(os.listdir(), columns=['items'])

    dir_list_dimension = dir_list.loc[dir_list['items'].str.contains('Rapid Raw Data_Latest.xlsb.crdownload')].shape

    while True:
        os.chdir(r'C:\Users\i024605\Downloads')

        dir_list = pd.DataFrame(os.listdir(), columns=['items'])

        dir_list_dimension = dir_list.loc[dir_list['items'].str.contains('Rapid Raw Data_Latest.xlsb.crdownload')].shape

        print("No File Found yet")
        
        if dir_list_dimension[0] < 1:
            break
        
        else:
            pass
        
        sleep(2)
        
    print("File Found")

    driver.close()



# %% [markdown]
# ## Start importing needed file

# %%
## Initiate Webscrap

WebScrap()

## Perform Data Loading

# directory = r'C:\Users\I024605\Downloads'
directory = os.path.join(os.path.expanduser("~"), "Downloads")

list_of_files = glob.glob(directory+'\\Rapid Raw Data_Latest*.xlsb')
print(list_of_files)

path_holder = os.path.join(os.path.expanduser("~"), "Downloads\PProtect_Data")


for i in range (5):
    list_of_files = glob.glob(directory+'\\Rapid Raw Data_Latest*.xlsb')
    try:
        list_of_file = list_of_files[0]
    except:
        list_of_file = [0]
    print(i)
    print(list_of_files)
    if ('Rapid Raw Data_Latest' not in list_of_file) & (i<=3):
        sleep(2)
        print("1st Cond")
    elif ('Rapid Raw Data_Latest' not in list_of_file) & (i >3):
        ("Print 2nd Cond")
        WebScrap()
    elif 'Rapid Raw Data_Latest' in list_of_file:
        break
    else:
        pass
    
os.listdir(path_holder)

if 'PProtect_Data' in os.listdir(directory):
    shutil.rmtree(path_holder)
    time.sleep(1)
    os.mkdir(path_holder)
else:
    os.mkdir(path_holder)


shutil.move(list_of_files[0], path_holder)


dtype= {
    'PolicyNumber' : str, 
    'Status':str, 
    'PolicyPremium':float,
    'Agent Code':str, 
    'Agent Name':str, 
    'Branch Code':str,
    'referrer code':str,
}

df_protect1 = pd.read_excel(path_holder+"\\Rapid Raw Data_Latest.xlsb", sheet_name='Rapid Raw', dtype=dtype)

# df_protect1 = pd.read_excel(r"W:\Data Extracts\Rapid_Raw\Rapid_Raw Data (Latest Update).xlsb", sheet_name='Rapid Raw', dtype=dtype)

df_protect1


# %% [markdown]
# ## Date Tranfrom Function

# %%
def date_conv(df):
    df['datetime'] = pd.to_datetime(df['Effective date'], errors='coerce')
    m = pd.to_numeric(df['Effective date'], errors='coerce').notna()
    df.loc[m, 'datetime'] = pd.to_datetime(df['Effective date'][m].astype(float), errors='coerce', unit='D', origin='1899-12-30')
    df['Effective date'] = df['datetime']
    df.drop(['datetime'], axis=1, inplace=True)

    return df

# %% [markdown]
# ## Trimming Function

# %%
def strip_element(x):
    if isinstance(x, str):
        return x.strip()
    return 

# %% [markdown]
# ## Database Function

# %%
def sql_connection(sql_query, dbase):
    # Replace with your actual SQL Server connection details
    server = 'PPBWDLC0SG7A1'
    database = dbase
    username = 'admin'
    password = 'Openlab@123'

    # Define the connection string
    connection_string = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
    )


    try:
        # Establish the connection
        conn = pyodbc.connect(connection_string)

        # Define the SELECT query
        query = sql_query

        # Execute the query and fetch data into a DataFrame
        df = pd.read_sql(query, conn)

        # Execute Trims 
        df = df.applymap(strip_element)

        # Print the data
        print(df)


    except pyodbc.Error as e:
        print("Error connecting to SQL Server:", e)
    finally:
        # Close the connection
        conn.close()

    return df

# %% [markdown]
# ## Getting the required columns only

# %%
df_protect2 = df_protect1.copy()

# Getting the needed data
df_protect2 = df_protect2[['PolicyNumber', 'Status', 'Effective date', 'PolicyPremium', 'Agent Code', 'Agent Name', 'Branch Code', 'referrer code', 'Campaign code']].copy()
df_protect2_tmp = df_protect2[['PolicyNumber', 'Status', 'Effective date', 'Agent Code', 'Agent Name', 'Branch Code', 'referrer code', 'Campaign code']].copy()

# Removing Trim
df_protect2_tmp = df_protect2_tmp.applymap(strip_element)
df_protect2_tmp['Effective date'] = df_protect2['Effective date']
df_protect2_tmp['PolicyPremium'] = df_protect2['PolicyPremium']
df_protect2 = df_protect2_tmp
df_protect2 

# Adjusting the Date
df_protect2_date = date_conv(df_protect2)
df_protect2_date = df_protect2_date.loc[df_protect2_date['Effective date'].dt.strftime("%Y")=='2024']
df_protect2_date

# %% [markdown]
# ## Fixing the Columns Status and Referrer Code

# %%
df_protect2_date['Status'] = np.where(df_protect2_date['Status']=='0x2a', "Reject", df_protect2_date['Status'])
df_protect2_date['referrer code'] = np.where(df_protect2_date['referrer code']=='0x2a', "", df_protect2_date['referrer code'])
df_protect2 = df_protect2_date.copy()
df_protect2

# %% [markdown]
# ## ANP

# %%
df_protect3 = df_protect2.copy()

df_protect3['ANP'] = df_protect3['PolicyPremium'].mul(12.0)

df_protect3

# %% [markdown]
# ## Case Count

# %%
df_protect4 = df_protect3.copy()

df_protect4['Case Count'] = 1

df_protect4

# %% [markdown]
# ## Old Load MasterFile

# %%
# query = '''
#     WITH MfileoftheMonth AS 
#     ( 
# 		Select 
# 		[RELA CODE]
#       ,[VALIDATION CODE]
#       ,[BANK  CODE]
#       ,[BPI-AIA  BRANCH CODE]
#       ,[BRANCH CODE  FOR BPI-AIA]
#       ,[CLUSTER CODE]
#       ,[BANK]
#       ,[BRANCH NAME]
#       ,[DIVISION]
#       ,[AREA]
#       ,[BRANCH TYPE]
#       ,[BUSINESS  MANAGER CODE]
#       ,[BUSINESS MANAGER]
#       ,[ASST BUSINESS MANAGER]
#       ,[BUSINESS DIRECTOR]
#       ,[BPI DIVISION HEAD]
#       ,[RM REF CODE]
#       ,[RM]
#       ,[SA]
#       ,[SO]
#       ,[BRANCH_ADD]
#       ,[BRANCH_TEL]
#       ,[TEL 2]
#       ,[TEL 3]
#       ,[BRANCH_FAX]
#       ,[ZIP]
#       ,[TERRITORY  CODE]
#       ,[TERRITORY]
#       ,[BPI-AIA  DIVISION CODE]
#       ,[BPI-AIA DIVISION]
#       ,[BPI-AIA  AREA CODE]
#       ,[BPI-AIA AREA]
#       ,[PERSONAL BSE 1]
#       ,[PER1 CODE]
#       ,[PERSONAL BSE 2]
#       ,[PER2 CODE]
#       ,[PERSONAL BSE 3]
#       ,[PER3 CODE]
#       ,[PERSONAL BSE 4]
#       ,[PER4 CODE]
#       ,[PERSONAL BSE 5]
#       ,[PER5 CODE]
#       ,[PREFERRED BSE 1]
#       ,[PRF1 CODE]
#       ,[PREFERRED BSE 2]
#       ,[PRF2 CODE]
#       ,[PREFERRED BSE 3]
#       ,[PRF3 CODE]
#       ,[BAM]
#       ,[BAM CODE]
#       ,[BDM]
#       ,[BDM CODE]
#       ,[TSH]
#       ,[TSH CODE]
#       ,[BRANCH OLD NAME]
#       ,[STATUS OF BRANCHES BASED ON BSC HIERARCHY]
#       ,[WWDate]
# 		, mlu.[Index]  FROM db_MasterFile_2024.[dbo].tbl_MasterFile_2024 AS Mfile24 
# 		LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(Mfile24.WWDate,3) 
# 	)
# 	,
# 	RMfileoftheMonth AS
# 	(
# 		Select 
# 		   [RELA CODE] = NULL
# 		  ,[VALIDATION CODE] = NULL
# 		  ,[BANK  CODE] = NULL
# 		  ,[BPI-AIA  BRANCH CODE]
# 		  ,[BRANCH CODE  FOR BPI-AIA]
# 		  ,[CLUSTER CODE] = NULL
# 		  ,[BANK]
# 		  ,[BRANCH NAME]
# 		  ,[ R BANK DIVISION] AS [DIVISION]
# 		  ,[R BANK AREA] AS [AREA]
# 		  ,[BRANCH TYPE]
# 		  ,[BUSINESS  MANAGER CODE] = NULL
# 		  ,[BUSINESS MANAGER] = NULL
# 		  ,[ASST BUSINESS MANAGER] = NULL
# 		  ,[BUSINESS DIRECTOR] 
# 		  ,[BPI DIVISION HEAD]
# 		  ,[RM REF CODE] = NULL
# 		  ,[RM] = NULL
# 		  ,[SA] = NULL
# 		  ,[SO] = NULL
# 		  ,[BRANCH_ADD] = NULL
# 		  ,[BRANCH_TEL] = NULL
# 		  ,[TEL 2] = NULL
# 		  ,[TEL 3] = NULL
# 		  ,[BRANCH_FAX] = NULL
# 		  ,[ZIP] = NULL
# 		  ,[TERRITORY  CODE]
# 		  ,[TERRITORY]
# 		  ,[BPI-AIA  DIVISION CODE]
# 		  ,[BPI-AIA DIVISION]
# 		  ,[BPI-AIA  AREA CODE]
# 		  ,[BPI-AIA AREA]
# 		  ,[BSE 1] AS [PERSONAL BSE 1]
# 		  ,[BSE1 CODE] AS [PER1 CODE]
# 		  ,[BSE 2] AS [PERSONAL BSE 2]
# 		  ,[BSE2 CODE] AS [PER2 CODE]
# 		  ,[BSE 3] AS [PERSONAL BSE 3]
# 		  ,[PER3 CODE] = NULL
# 		  ,[PERSONAL BSE 4] = NULL
# 		  ,[PER4 CODE] = NULL
# 		  ,[PERSONAL BSE 5] = NULL
# 		  ,[PER5 CODE] = NULL
# 		  ,[PREFERRED BSE 1] = NULL
# 		  ,[PRF1 CODE] = NULL
# 		  ,[PREFERRED BSE 2] = NULL
# 		  ,[PRF2 CODE] = NULL
# 		  ,[PREFERRED BSE 3] = NULL
# 		  ,[PRF3 CODE] = NULL
# 		  ,[BAM]
# 		  ,[BAM CODE]
# 		  ,[BDM]
# 		  ,[BDM CODE]
# 		  ,[TSH] = NULL
# 		  ,[TSH CODE] = NULL
# 		  ,[BRANCH OLD NAME] = NULL
# 		  ,[STATUS OF BRANCHES BASED ON BSC HIERARCHY] = NULL
# 		  ,[WWDate]
# 		,mlu.[Index]  FROM db_MasterFile_2024.[dbo].tbl_RBankMasterFile_2024 AS RMfile24 
# 		LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(RMfile24.WWDate,3) 
# 		) 

#     SELECT MFile.* FROM MfileoftheMonth AS MFile WHERE MFile.[INDEX] = (SELECT MAX(mlu.[INDEX]) FROM MfileoftheMonth mlu)
# 	UNION ALL
# 	SELECT RMFile.* FROM RMfileoftheMonth AS RMFile WHERE RMFile.[INDEX] = (SELECT MAX(mlu.[INDEX]) FROM RMfileoftheMonth mlu)
# '''
# db_masterfile = sql_connection(query, 'db_MasterFile_2024')
# db_masterfile

# %% [markdown]
# ## New Load Masterfile

# %%
query = '''
    WITH MfileoftheMonth AS 
    ( 
	Select 
	   [RELA CODE]
      ,[VALIDATION CODE]
      ,[BANK  CODE]
      ,[BPI-AIA  BRANCH CODE]
      ,[BRANCH CODE  FOR BPI-AIA]
      ,[CLUSTER CODE]
      ,[BANK]
      ,[BRANCH NAME]
      ,[DIVISION]
      ,[AREA]
      ,[BRANCH TYPE]
      ,[BUSINESS  MANAGER CODE]
      ,[BUSINESS MANAGER]
      ,[ASST BUSINESS MANAGER]
      ,[BUSINESS DIRECTOR]
      ,[BPI DIVISION HEAD]
      ,[RM REF CODE]
      ,[RM]
      ,[SA]
      ,[SO]
      ,[BRANCH_ADD]
      ,[BRANCH_TEL]
      ,[TEL 2]
      ,[TEL 3]
      ,[BRANCH_FAX]
      ,[ZIP]
      ,[TERRITORY  CODE]
      ,[TERRITORY]
      ,[BPI-AIA  DIVISION CODE]
      ,[BPI-AIA DIVISION]
      ,[BPI-AIA  AREA CODE]
      ,[BPI-AIA AREA]
      ,[PERSONAL BSE 1]
      ,[PER1 CODE]
      ,[PERSONAL BSE 2]
      ,[PER2 CODE]
      ,[PERSONAL BSE 3]
      ,[PER3 CODE]
      ,[PERSONAL BSE 4]
      ,[PER4 CODE]
      ,[PERSONAL BSE 5]
      ,[PER5 CODE]
      ,[PREFERRED BSE 1]
      ,[PRF1 CODE]
      ,[PREFERRED BSE 2]
      ,[PRF2 CODE]
      ,[PREFERRED BSE 3]
      ,[PRF3 CODE]
      ,[BAM]
      ,[BAM CODE]
      ,[BDM]
      ,[BDM CODE]
      ,[TSH]
      ,[TSH CODE]
      ,[BRANCH OLD NAME]
      ,[STATUS OF BRANCHES BASED ON BSC HIERARCHY]
      ,[WWDate]
	FROM db_MasterFile_2024.[dbo].tbl_MasterFile_2024 AS Mfile24 
	)
	,
	RMfileoftheMonth AS
	(
		Select 
		   [RELA CODE] = NULL
		  ,[VALIDATION CODE] = NULL
		  ,[BANK  CODE] = NULL
		  ,RMfile24.[BPI-AIA  BRANCH CODE]
		  ,RMfile24.[BRANCH CODE  FOR BPI-AIA]
		  ,[CLUSTER CODE] = NULL
		  ,RMfile24.[BANK]
		  ,RMfile24.[BRANCH NAME]
		  ,RMfile24.[ R BANK DIVISION] AS [DIVISION]
		  ,RMfile24.[R BANK AREA] AS [AREA]
		  ,RMfile24.[BRANCH TYPE]
		  ,[BUSINESS  MANAGER CODE] = NULL
		  ,[BUSINESS MANAGER] = NULL
		  ,[ASST BUSINESS MANAGER] = NULL
		  ,RMfile24.[BUSINESS DIRECTOR] 
		  ,RMfile24.[BPI DIVISION HEAD]
		  ,[RM REF CODE] = NULL
		  ,[RM] = NULL
		  ,[SA] = NULL
		  ,[SO] = NULL
		  ,[BRANCH_ADD] = NULL
		  ,[BRANCH_TEL] = NULL
		  ,[TEL 2] = NULL
		  ,[TEL 3] = NULL
		  ,[BRANCH_FAX] = NULL
		  ,[ZIP] = NULL
		  ,RMfile24.[TERRITORY  CODE]
		  ,RMfile24.[TERRITORY]
		  ,RMfile24.[BPI-AIA  DIVISION CODE]
		  ,RMfile24.[BPI-AIA DIVISION]
		  ,RMfile24.[BPI-AIA  AREA CODE]
		  ,RMfile24.[BPI-AIA AREA]
		  ,RMfile24.[BSE 1] AS [PERSONAL BSE 1]
		  ,RMfile24.[BSE1 CODE] AS [PER1 CODE]
		  ,RMfile24.[BSE 2] AS [PERSONAL BSE 2]
		  ,RMfile24.[BSE2 CODE] AS [PER2 CODE]
		  ,RMfile24.[BSE 3] AS [PERSONAL BSE 3]
		  ,[PER3 CODE] = NULL
		  ,[PERSONAL BSE 4] = NULL
		  ,[PER4 CODE] = NULL
		  ,[PERSONAL BSE 5] = NULL
		  ,[PER5 CODE] = NULL
		  ,[PREFERRED BSE 1] = NULL
		  ,[PRF1 CODE] = NULL
		  ,[PREFERRED BSE 2] = NULL
		  ,[PRF2 CODE] = NULL
		  ,[PREFERRED BSE 3] = NULL
		  ,[PRF3 CODE] = NULL
		  ,RMfile24.[BAM]
		  ,RMfile24.[BAM CODE]
		  ,RMfile24.[BDM]
		  ,RMfile24.[BDM CODE]
		  ,[TSH] = NULL
		  ,[TSH CODE] = NULL
		  ,[BRANCH OLD NAME] = NULL
		  ,[STATUS OF BRANCHES BASED ON BSC HIERARCHY] = NULL
		  ,RMfile24.[WWDate]
		FROM db_MasterFile_2024.[dbo].tbl_RBankMasterFile_2024 AS RMfile24 
		),
	Mfile_RMFile AS
	(
		SELECT MFile.* FROM MfileoftheMonth AS MFile 
		UNION 
		SELECT RMFile.* FROM RMfileoftheMonth AS RMFile 
	)
	Select MRFile2024.*, mlu.[INDEX]
	,MRFile2024.[BRANCH CODE  FOR BPI-AIA] + '_' + CAST(mlu.[INDEX] AS nvarchar) + '_' +RIGHT(RTRIM(MRFile2024.[WWDate]), 4) AS [BRANCH CODE  FOR BPI-AIA DATE]

	FROM Mfile_RMFile AS MRFile2024 
	Left JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(MRFile2024.WWDate,3)
	ORDER BY mlu.[INDEX]
'''

db_masterfile = sql_connection(query, 'db_MasterFile_2024')
db_masterfile

# %% [markdown]
# ## Old Load BSE MasterFile

# %%
# query = '''
# WITH CorpsoloftheMonth AS
# (
# SELECT * FROM db_MasterFile_2024.[dbo].tbl_CorpsolMasterFile_2024 AS COR24
# LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(COR24.WWDate,3) 
# ),
# CORPS24 AS
# (
# SELECT COR24.* FROM CorpsoloftheMonth AS COR24 WHERE COR24.[INDEX] = (SELECT MAX(cor.[INDEX]) FROM CorpsoloftheMonth cor)
# ),
# BSEoftheMonth AS 
# ( 
# Select BSE24.*, mlu.[Index]  FROM db_MasterFile_2024.[dbo].tbl_BSEMasterFile_2024 AS BSE24 
# LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(BSE24.WWDate,3) 
# ),
# BSES24 AS
# (
# SELECT BSE24.* FROM BSEoftheMonth AS BSE24 WHERE BSE24.[INDEX] = (SELECT MAX(bse.[INDEX]) FROM BSEoftheMonth bse)
# ),
# BSECORP AS
# (
# SELECT 
#        BSES24.[POSITION]
#       ,BSES24.[EE CODE]
#       ,BSES24.[BSE FULL NAME]
#       ,BSES24.[BSE CODE]
#       ,BSES24.[SURNAME]
#       ,BSES24.[FIRST NAME]
#       ,BSES24.[MIDDLE NAME]
#       ,BSES24.[SEGMENT]
#       ,BSES24.[HYBRID]
#       ,BSES24.[JUNIOR/ SENIOR]
#       ,BSES24.[OLD/NEW]
#       ,BSES24.[TENURE  (CURRENT POSITION)]
#       ,BSES24.[DEPLOYMENT DATE (MM/DD/YYYY)]
#       ,BSES24.[DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)]
#       ,BSES24.[MOBILE NUMBERS]
#       ,BSES24.[EMAILADD]
#       ,CASE
# 		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
# 		ELSE BSES24.[BPI-AIA AREA]
# 	   END AS [BPI-AIA AREA]
#       ,CASE
# 		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
# 		ELSE BSES24.[BPI-AIA DIVISION]
# 	   END AS [BPI-AIA DIVISION]
#       ,CASE
# 		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
# 		ELSE BSES24.[TERRITORY]
# 	   END AS [TERRITORY]
#       ,BSES24.[BAM]
#       ,BSES24.[BAM CODE]
#       ,BSES24.[BDM]
#       ,BSES24.[BDM CODE]
#       ,BSES24.[TSH]
#       ,BSES24.[TSH CODE]
#       ,BSES24.[TL NAME]
#       ,BSES24.[TL CODE]
#       ,BSES24.[WWDate]
#       ,BSES24.[BATCH]
# FROM BSES24

# UNION ALL

# SELECT 
# 	 CORPS24.[POSITION]	
# 	,CORPS24.[EE CODE]
# 	,CORPS24.[FULL NAME]
# 	,CORPS24.[POSITION CODE]
# 	,CORPS24.[SURNAME]	
# 	,CORPS24.[FIRST NAME]	
# 	,CORPS24.[MIDDLE NAME]
# 	,CORPS24.[SEGMENT]
# 	,'Hybrid' = NULL
# 	,'JUNIOR/ SENIOR' = NULL
# 	,CORPS24.[OLD/NEW]
# 	,CORPS24.[TENURE  (CURRENT POSITION)]
# 	,CORPS24.[DEPLOYMENT DATE]
# 	,'DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)' = NULL
# 	,CORPS24.[MOBILE NUMBERS]
# 	,CORPS24.[EMAILADD]
# 	,CASE
# 		WHEN [POSITION] IN ('MBBSA', 'CBSS') THEN 'CORPSOL'
# 		ELSE CORPS24.[BPLAC AREA]
# 	END AS [BPLAC AREA]
# 	,CASE
# 		WHEN [POSITION] IN ('MBBSA', 'CBSS') THEN 'CORPSOL'
# 		ELSE CORPS24.[BPLAC DIV]
# 	END AS [BPLAC DIV]
# 	,'TERRITORY'= 'CORPSOL'
# 	,CORPS24.[CSM] AS [BAM]
# 	,CORPS24.[CSM CODE] AS [BAM CODE]
# 	,CORPS24.[HEAD] AS [BDM]
# 	,CORPS24.[CHANNEL HEAD CODE] AS [BDM CODE]
# 	,'TSH' = NULL
# 	,'TSH CODE' = NULL
# 	,'TL NAME' = NULL
# 	,'TL CODE' = NULL
# 	,CORPS24.[WWDate]
# 	,'BATCH' = NULL
# FROM CORPS24
# WHERE [POSITION] IN ('MBBSA', 'CBSS')
# )
# SELECT * FROM BSECORP
# '''
# db_bse_masterfile = sql_connection(query, 'db_MasterFile_2024')
# db_bse_masterfile

# %% [markdown]
# ## New Load BSE MasterFile

# %%
query = '''
WITH CorpsoloftheMonth AS
(
SELECT * FROM db_MasterFile_2024.[dbo].tbl_CorpsolMasterFile_2024 AS COR24
LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(COR24.WWDate,3) 
),
CORPS24 AS
(
SELECT COR24.* FROM CorpsoloftheMonth AS COR24 
),
BSEoftheMonth AS 
( 
Select BSE24.*, mlu.[Index]  FROM db_MasterFile_2024.[dbo].tbl_BSEMasterFile_2024 AS BSE24 
LEFT JOIN [db_Lookup].[dbo].tbl_monthindex_unique AS mlu ON mlu.[MON] = left(BSE24.WWDate,3) 
),
BSES24 AS
(
SELECT BSE24.* FROM BSEoftheMonth AS BSE24 
),
BSECORP AS
(
SELECT 
       BSES24.[POSITION]
      ,BSES24.[EE CODE]
      ,BSES24.[BSE FULL NAME]
      ,BSES24.[BSE CODE]
      ,BSES24.[SURNAME]
      ,BSES24.[FIRST NAME]
      ,BSES24.[MIDDLE NAME]
      ,BSES24.[SEGMENT]
      ,BSES24.[HYBRID]
      ,BSES24.[JUNIOR/ SENIOR]
      ,BSES24.[OLD/NEW]
      ,BSES24.[TENURE  (CURRENT POSITION)]
      ,BSES24.[DEPLOYMENT DATE (MM/DD/YYYY)]
      ,BSES24.[DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)]
      ,BSES24.[MOBILE NUMBERS]
      ,BSES24.[EMAILADD]
      ,CASE
		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
		ELSE BSES24.[BPI-AIA AREA]
	   END AS [BPI-AIA AREA]
      ,CASE
		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
		ELSE BSES24.[BPI-AIA DIVISION]
	   END AS [BPI-AIA DIVISION]
      ,CASE
		WHEN [POSITION] = 'ABBSE' THEN 'ABG'
		ELSE BSES24.[TERRITORY]
	   END AS [TERRITORY]
      ,BSES24.[BAM]
      ,BSES24.[BAM CODE]
      ,BSES24.[BDM]
      ,BSES24.[BDM CODE]
      ,BSES24.[TSH]
      ,BSES24.[TSH CODE]
      ,BSES24.[TL NAME]
      ,BSES24.[TL CODE]
      ,BSES24.[WWDate]
      ,BSES24.[BATCH]
	  ,BSES24.[INDEX]
FROM BSES24

UNION ALL

SELECT 
	 CORPS24.[POSITION]	
	,CORPS24.[EE CODE]
	,CORPS24.[FULL NAME]
	,CORPS24.[POSITION CODE]
	,CORPS24.[SURNAME]	
	,CORPS24.[FIRST NAME]	
	,CORPS24.[MIDDLE NAME]
	,CORPS24.[SEGMENT]
	,'Hybrid' = NULL
	,'JUNIOR/ SENIOR' = NULL
	,CORPS24.[OLD/NEW]
	,CORPS24.[TENURE  (CURRENT POSITION)]
	,CORPS24.[DEPLOYMENT DATE]
	,'DEPLOYMENT DATE (CURRENT POSITION) (MM/DD/YYYY)' = NULL
	,CORPS24.[MOBILE NUMBERS]
	,CORPS24.[EMAILADD]
	,CASE
		WHEN [POSITION] IN ('MBBSA', 'CBSS') THEN 'CORPSOL'
		ELSE CORPS24.[BPLAC AREA]
	END AS [BPLAC AREA]
	,CASE
		WHEN [POSITION] IN ('MBBSA', 'CBSS') THEN 'CORPSOL'
		ELSE CORPS24.[BPLAC DIV]
	END AS [BPLAC DIV]
	,'TERRITORY'= 'CORPSOL'
	,CORPS24.[CSM] AS [BAM]
	,CORPS24.[CSM CODE] AS [BAM CODE]
	,CORPS24.[HEAD] AS [BDM]
	,CORPS24.[CHANNEL HEAD CODE] AS [BDM CODE]
	,'TSH' = NULL
	,'TSH CODE' = NULL
	,'TL NAME' = NULL
	,'TL CODE' = NULL
	,CORPS24.[WWDate]
	,'BATCH' = NULL
	,CORPS24.[INDEX]
FROM CORPS24
WHERE [POSITION] IN ('MBBSA', 'CBSS')
)
SELECT BSECORP.*
		,BSECORP.[BSE CODE]+'_'+CAST(BSECORP.[INDEX] AS nvarchar)+'_'+RIGHT(RTRIM(BSECORP.[WWDate]), 4) AS [BSE CODE DATE] FROM BSECORP 

'''
db_bse_masterfile = sql_connection(query, 'db_MasterFile_2024')
db_bse_masterfile

# %% [markdown]
# ## Megrge with BSE Database via BSE Code

# %%
df_protect4['Agent Code'] + '_' + df_protect4['Effective date'].dt.strftime('%#m') + '_' + df_protect4['Effective date'].dt.strftime('%Y')

# %%
df_protect5 = df_protect4.copy()

df_protect5['BSE CODE DATE'] = df_protect5['Agent Code'] + '_' + df_protect5['Effective date'].dt.strftime('%#m') + '_' + df_protect5['Effective date'].dt.strftime('%Y')

db_bse_master_temp = db_bse_masterfile[['BSE CODE', 'SEGMENT', 'BPI-AIA AREA', 'BPI-AIA DIVISION', 'TERRITORY', 'BAM', 'BAM CODE', 'BDM', 'BDM CODE', 'TSH', 'TSH CODE', 'BSE CODE DATE']].copy()
db_bse_master_temp = db_bse_master_temp.drop_duplicates(subset=['BSE CODE DATE']) # Remove duplicates so lookup merge only returns first match

df_protect5['BSE CODE DATE'] = df_protect5['BSE CODE DATE'].str.strip()
df_merge = df_protect5.merge(db_bse_master_temp, left_on=['BSE CODE DATE'], right_on=['BSE CODE DATE'], how='left', suffixes=["", '_db_bse_master_temp'])

df_protect5 = df_merge.copy()

df_protect5.drop('BSE CODE DATE', axis=1, inplace=True)
df_protect5


# %% [markdown]
# ## Merge With MasterFile via Branch Name

# %%
df_protect6 = df_protect5.copy()

df_protect6['BRANCH CODE  FOR BPI-AIA DATE'] = df_protect6['Branch Code'] + '_' + df_protect6['Effective date'].dt.strftime('%#m') + '_' + df_protect6['Effective date'].dt.strftime('%Y')

db_master_temp = db_masterfile[['BRANCH CODE  FOR BPI-AIA', 'BRANCH NAME', 'DIVISION',	'AREA', 'TERRITORY', 'BPI-AIA DIVISION', 'BPI-AIA AREA', 'BAM', 'BAM CODE', 'BDM', 'BDM CODE', 'TSH', 'TSH CODE', 'BRANCH CODE  FOR BPI-AIA DATE']].copy()
db_master_temp = db_master_temp.drop_duplicates(subset=['BRANCH CODE  FOR BPI-AIA DATE']) # Remove duplicates so lookup merge only returns first match

df_protect6['BRANCH CODE  FOR BPI-AIA DATE'] = df_protect6['BRANCH CODE  FOR BPI-AIA DATE'].str.strip()
df_merge = df_protect6.merge(db_master_temp, left_on=['BRANCH CODE  FOR BPI-AIA DATE'], right_on=['BRANCH CODE  FOR BPI-AIA DATE'], how='left', suffixes=["", '_db_master_temp'])

df_protect6 = df_merge.copy()

del(db_master_temp)

df_protect6.drop('BRANCH CODE  FOR BPI-AIA DATE', axis=1)

df_protect6

# %%
df_protect6.loc[df_protect6['Agent Code'] == '090212359']
df_protect6.loc[df_protect6['PolicyNumber'] == '5101234876']

# %% [markdown]
# ## Matching Conditions

# %%
df_protect7 = df_protect6.copy()


df_protect7['BPI-AIA DIVISION'] = np.where(df_protect7['BPI-AIA DIVISION'].isna(),  df_protect7['BPI-AIA DIVISION_db_master_temp'], df_protect7['BPI-AIA DIVISION'])

df_protect7['BPI-AIA AREA'] = np.where(df_protect7['BPI-AIA AREA'].isna(),  df_protect7['BPI-AIA AREA_db_master_temp'], df_protect7['BPI-AIA AREA'])

df_protect7['TERRITORY'] = np.where(df_protect7['TERRITORY'].isna(), df_protect7['TERRITORY_db_master_temp'], df_protect7['TERRITORY'])

df_protect7['BDM'] = np.where(df_protect7['BDM'].isna(), df_protect7['BDM_db_master_temp'], df_protect7['BDM'])

df_protect7['BDM CODE'] = np.where(df_protect7['BDM CODE'].isna(), df_protect7['BDM CODE_db_master_temp'], df_protect7['BDM CODE'])

df_protect7['BAM'] = np.where(df_protect7['BAM'].isna(), df_protect7['BAM_db_master_temp'], df_protect7['BAM'])

df_protect7['BAM CODE'] = np.where(df_protect7['BAM CODE'].isna(), df_protect7['BAM CODE_db_master_temp'], df_protect7['BAM CODE'])

df_protect7['TSH'] = np.where(df_protect7['TSH'].isna(), df_protect7['TSH_db_master_temp'], df_protect7['TSH'])

df_protect7['TSH CODE'] = np.where(df_protect7['TSH CODE'].isna(), df_protect7['TSH CODE_db_master_temp'], df_protect7['TSH CODE'])


## Matching Branch Code
df_protect7['BRANCH NAME'] = np.where(((df_protect7['Branch Code'].notna()) & (df_protect7['BRANCH NAME'].isna())), 'SG', df_protect7['BRANCH NAME'] )

df_protect7['DIVISION'] = np.where(((df_protect7['Branch Code'].notna()) & (df_protect7['DIVISION'].isna())), 'SG', df_protect7['DIVISION'] )

df_protect7['AREA'] = np.where(((df_protect7['Branch Code'].notna()) & (df_protect7['AREA'].isna())), 'SG', df_protect7['AREA'] )

df_protect7['BPI-AIA DIVISION'] = np.where(((df_protect7['Branch Code'].notna()) & (df_protect7['BPI-AIA DIVISION'].isna())), 'SG', df_protect7['BPI-AIA DIVISION'] )

df_protect7['BPI-AIA AREA'] = np.where(((df_protect7['Branch Code'].notna()) & (df_protect7['BPI-AIA AREA'].isna())), 'SG', df_protect7['BPI-AIA AREA'] )

df_protect7['TERRITORY'] = np.where(((df_protect7['Branch Code'].notna()) & (df_protect7['TERRITORY'].isna())), 'SG', df_protect7['TERRITORY'] )


## Matching Agent Code - If Agent Code Have a Value and the Corresponding Partner is Blank then its SG

df_protect7['BDM'] = np.where(((df_protect7['Agent Code'].notna()) & (df_protect7['BDM'].isna())), "SG", df_protect7['BDM'])

df_protect7['BDM CODE'] = np.where(((df_protect7['Agent Code'].notna()) & (df_protect7['BDM CODE'].isna())), "SG", df_protect7['BDM CODE'])

df_protect7['BAM'] = np.where(((df_protect7['Agent Code'].notna()) & (df_protect7['BAM'].isna())), "SG", df_protect7['BAM'])

df_protect7['BAM CODE'] = np.where(((df_protect7['Agent Code'].notna()) & (df_protect7['BAM CODE'].isna())), "SG", df_protect7['BAM CODE'])

df_protect7['TSH'] = np.where(((df_protect7['Agent Code'].notna()) & (df_protect7['TSH'].isna())), "SG", df_protect7['TSH'])

df_protect7['TSH CODE'] = np.where(((df_protect7['Agent Code'].notna()) & (df_protect7['TSH CODE'].isna())), "SG", df_protect7['TSH CODE'])

df_protect7['COUNT'] = 1

df_protect7

# %% [markdown]
# ## Getting the Require Attributes Only

# %%
df_protect8 = df_protect7[['PolicyNumber', 'Campaign code', 'Status', 'Effective date', 'PolicyPremium', 'Agent Code', 'Agent Name',
       'Branch Code', 'referrer code',   'ANP',
       'Case Count', 'BRANCH NAME', 'DIVISION',
       'AREA', 'TERRITORY', 'BPI-AIA DIVISION', 'BPI-AIA AREA', 'BDM CODE', 'BDM', 
       'BAM CODE',  'BAM',  'TSH CODE', 'TSH',  'SEGMENT', 'COUNT']].copy()

df_protect8

# %%
df_protect2['Status'].value_counts()

# %% [markdown]
# ## Adding Status Attributes

# %%
df_protect9 = df_protect8.copy()

# condition = [df_protect9['Status'].str.lower() == "inforce", df_protect9['Status'].str.lower() == "reinstate", df_protect9['Status'].str.lower() == "reject", df_protect9['Status'].str.lower() == "cancel", df_protect9['Status'].str.lower() == "lapse"]
# choicelist = ["Approved", "Approved", "Rejected" ,"Cancel", "Approved"]

# df_protect9['APPROVAL STATUS'] = np.select(condition, choicelist, "Pending")

condition = [df_protect9['Status'].str.lower() == "reject", df_protect9['Status'].str.lower() == "cancel"]
choicelist = ["Not Approved", "Not Approved"]

df_protect9['APPROVAL STATUS'] = np.select(condition, choicelist, "Approved")

df_protect9['APPROVAL STATUS'] = np.where(df_protect9['Status'].str.lower().str.contains("terminate"), "Not Approved", df_protect9['APPROVAL STATUS'])

df_protect9


# %% [markdown]
# ## Adding FTD Attributes

# %%
yestreday = (datetime.now() + timedelta(days=-1)).strftime("%Y-%m-%d")
yestreday

df_protect10 = df_protect9.copy()

df_protect10['FTD'] = np.where(df_protect10['Effective date']==yestreday, "FTD", "Not FTD")

df_protect10

# %% [markdown]
# ## Adding RBank in the End

# %%
df_protect11 = df_protect10.copy()
df_protect11["BANK CATEGORY"] = ''


df_protect11["BANK CATEGORY"] = 'BPI'

df_protect11["BANK CATEGORY"] = np.where(df_protect11['Branch Code'].str.startswith('5'), 'RBANK', df_protect11["BANK CATEGORY"])
df_protect11["BANK CATEGORY"] = np.where(df_protect11['SEGMENT'] == 'ABG', 'ABG', df_protect11["BANK CATEGORY"])
df_protect11["BANK CATEGORY"] = np.where((df_protect11['SEGMENT'] == 'CBG'), 'CORPSOL', df_protect11["BANK CATEGORY"])
df_protect11["BANK CATEGORY"] = np.where((df_protect11['SEGMENT'] == 'BBG'), 'CORPSOL', df_protect11["BANK CATEGORY"])
df_protect11["BANK CATEGORY"] = np.where((df_protect11['SEGMENT'] == 'IBG'), 'CORPSOL', df_protect11["BANK CATEGORY"])
df_protect11["BANK CATEGORY"] = np.where(((df_protect11["Branch Code"] == '040000009') | (df_protect11["Branch Code"] == '710000001')), "ABG", df_protect11["BANK CATEGORY"])

df_protect11["DIVISION"] = np.where((((df_protect11["Branch Code"] == '040000009') | (df_protect11["Branch Code"] == '710000001')) & ((df_protect11['SEGMENT'] == 'ABG'))), "ABG", df_protect11["DIVISION"])
df_protect11["AREA"] = np.where(((df_protect11["Branch Code"] == '040000009') | (df_protect11["Branch Code"] == '710000001') & (df_protect11['SEGMENT'] == 'ABG')), "ABG", df_protect11["AREA"])


# %% [markdown]
# ## Load BSE DMTM

# %%
query = '''
WITH SALES_T1 AS 
  (
	  SELECT	[AGENT ID] AS [BSE CODE],
				[AGENT NAME] AS [BSE NAME],
				[TEAM LEADER] AS [BAM NAME],
				[WWDate]
	  FROM [db_MasterFile_2024].[dbo].[tbl_DMTM_2024]
	  WHERE [STATUS] = 'TSR' and [TEAM] = 'SALES TEAM'),

BAM_T1 AS(
	  SELECT [AGENT ID] AS [BAM CODE], 
	  LEFT([AGENT NAME], CHARINDEX(' ', [AGENT NAME]) - 1) + RIGHT([AGENT NAME], LEN([AGENT NAME]) - CHARINDEX(' ', [AGENT NAME]) - 2) AS [BAM NAME],
	  TRIM(RIGHT([TEAM LEADER], CHARINDEX(' ', REVERSE([TEAM LEADER])) - 1)) AS [BDM NAME],
	  [AGENT NAME] AS [BAM FULL NAME]
	  FROM [db_MasterFile_2024].[dbo].[tbl_DMTM_2024]
	  WHERE [STATUS] = 'TL' and [TEAM] = 'SALES TEAM' 
	  GROUP BY [AGENT ID], [AGENT NAME], [TEAM LEADER] ),




BDM_T1 AS(
	  SELECT [AGENT ID] AS [BDM CODE], 
	  TRIM(RIGHT([AGENT NAME], CHARINDEX(' ', REVERSE([AGENT NAME])) - 1)) AS [BDM NAME],
	  [AGENT NAME] AS [BDM FULL NAME]
	  FROM [db_MasterFile_2024].[dbo].[tbl_DMTM_2024]
	  WHERE [STATUS] = 'TSM' and [TEAM] = 'SUPPORT TEAM' 
	  GROUP BY [AGENT ID], [AGENT NAME] )
	,


SALES_TABLE_FINAL AS
(
	SELECT SALES_T1.[BSE CODE], SALES_T1.[BSE NAME], BAM_T1.[BAM CODE], BAM_T1.[BAM FULL NAME], BDM_T1.[BDM CODE], BDM_T1.[BDM FULL NAME], [WWDate] FROM SALES_T1 
	LEFT JOIN BAM_T1 
	ON SALES_T1.[BAM NAME] = BAM_T1.[BAM NAME]
	LEFT JOIN BDM_T1
	ON BDM_T1.[BDM NAME] = BAM_T1.[BDM NAME]
),

REINSTATEMENT_T1 AS 
  (
	  SELECT	[AGENT ID] AS [BSE CODE],
				[AGENT NAME] AS [BSE NAME],
				[TEAM LEADER] AS [BAM NAME],
				[WWDate]
	  FROM [db_MasterFile_2024].[dbo].[tbl_DMTM_2024]
	  WHERE [STATUS] = 'TSR' and [TEAM] = 'REINSTATEMENT'),

REINSTATEMENT_T1_BAM_T1 AS(
	  SELECT [AGENT ID] AS [BAM CODE], 
	  LEFT([AGENT NAME], CHARINDEX(' ', [AGENT NAME]) - 1) + RIGHT([AGENT NAME], LEN([AGENT NAME]) - CHARINDEX(' ', [AGENT NAME]) - 2) AS [BAM NAME],
	  TRIM(RIGHT([TEAM LEADER], CHARINDEX(' ', REVERSE([TEAM LEADER])) - 1)) AS [BDM NAME],
	  [AGENT NAME] AS [BAM FULL NAME]
	  FROM [db_MasterFile_2024].[dbo].[tbl_DMTM_2024]
	  WHERE [STATUS] = 'TL' and [TEAM] = 'REINSTATEMENT'
	  GROUP BY [AGENT ID], [AGENT NAME], [TEAM LEADER] ),

REINSTATEMENT_T1_BDM_T1 AS(
	  SELECT [AGENT ID] AS [BDM CODE], 
	  TRIM(RIGHT([AGENT NAME], CHARINDEX(' ', REVERSE([AGENT NAME])) - 1)) AS [BDM NAME],
	  [AGENT NAME] AS [BDM FULL NAME]
	  FROM [db_MasterFile_2024].[dbo].[tbl_DMTM_2024]
	  WHERE [STATUS] = 'TSM' and [TEAM] = 'SUPPORT TEAM' 
	  GROUP BY [AGENT ID], [AGENT NAME] ),

REINSTATEMENT_TABLE_FINAL AS
(
SELECT REINSTATEMENT_T1.[BSE CODE], REINSTATEMENT_T1.[BSE NAME], REINSTATEMENT_T1_BAM_T1.[BAM CODE], REINSTATEMENT_T1_BAM_T1.[BAM FULL NAME], 
		REINSTATEMENT_T1_BDM_T1.[BDM CODE], REINSTATEMENT_T1_BDM_T1.[BDM FULL NAME], [WWDate] FROM REINSTATEMENT_T1 
LEFT JOIN REINSTATEMENT_T1_BAM_T1 
ON REINSTATEMENT_T1.[BAM NAME] = REINSTATEMENT_T1_BAM_T1.[BAM NAME]
LEFT JOIN REINSTATEMENT_T1_BDM_T1
ON REINSTATEMENT_T1_BDM_T1.[BDM NAME] = REINSTATEMENT_T1_BDM_T1.[BDM NAME]
),

FINAL_TABLE AS (
SELECT *, [TSH NAME]='JOHN WILMAR C. CUE', [TSH CODE] = '030401618', ([BSE CODE] +'_'+ TRY_CAST(MONTH(TRY_CAST(WWDate AS DATE)) AS nvarchar) +'_'+ TRY_CAST(YEAR(TRY_CAST(WWDate AS DATE)) AS nvarchar)) AS [BSE CODE DATE] FROM SALES_TABLE_FINAL
UNION
SELECT *, [TSH NAME]='JOHN WILMAR C. CUE', [TSH CODE] = '030401618', ([BSE CODE] +'_'+ TRY_CAST(MONTH(TRY_CAST(WWDate AS DATE)) AS nvarchar) +'_'+ TRY_CAST(YEAR(TRY_CAST(WWDate AS DATE)) AS nvarchar)) AS [BSE CODE DATE]  FROM REINSTATEMENT_TABLE_FINAL
)
SELECT * FROM FINAL_TABLE WHERE TRY_CAST([WWDate] AS Date) >= '2024-09-01'
'''

db_bse_dmtm = sql_connection(query, 'db_MasterFile_2024')

db_bse_dmtm_1 = db_bse_dmtm.copy()
db_bse_dmtm_1 = db_bse_dmtm_1[['BSE CODE', 'BSE NAME', 'BAM CODE', 'BAM FULL NAME', 'BDM CODE', 'BDM FULL NAME', 'BSE CODE DATE']]

db_bse_dmtm_1.loc[db_bse_dmtm_1['BSE CODE DATE'] == '088412067_10_2024']

# %%
dmtm_list = db_bse_dmtm_1['BSE CODE'].to_list()
df_protect11.loc[df_protect11['Agent Code'].isin(dmtm_list)].head()
print(len(df_protect11.loc[df_protect11['Agent Code'].isin(dmtm_list)]))


# %%
df_protect12 = df_protect11.copy()

df_protect12['BSE CODE DATE'] = df_protect12['Agent Code'] + '_' + df_protect12['Effective date'].dt.strftime('%#m') + '_' + df_protect12['Effective date'].dt.strftime('%Y')

db_bse_master_temp = db_bse_dmtm.copy()
db_bse_master_temp = db_bse_master_temp.drop_duplicates(subset=['BSE CODE DATE']) # Remove duplicates so lookup merge only returns first match

df_merge = df_protect12.merge(db_bse_master_temp, left_on=['BSE CODE DATE'], right_on=['BSE CODE DATE'], how='left', suffixes=["", '_db_bse_dmtm_temp'])

df_protect12 = df_merge.copy()

df_protect12['DMTM_TAG'] = ''
df_protect12.loc[df_protect12['BSE CODE'].notna(), 'DMTM_TAG'] = 'DMTM'

dmtm_list = df_protect12.loc[df_protect12['Agent Code'].isin(dmtm_list)]

df_protect12.drop(['BSE CODE DATE', 'BSE CODE'], axis=1, inplace=True)
df_protect12.head()


# %%
df_protect13 = df_protect12.copy()


df_protect13['BPI-AIA DIVISION'] = np.where(df_protect13['DMTM_TAG']=='DMTM',  "DMTM", df_protect13['BPI-AIA DIVISION'])

df_protect13['BPI-AIA AREA'] = np.where(df_protect13['DMTM_TAG']=='DMTM',  "DMTM", df_protect13['BPI-AIA AREA'])

df_protect13['TERRITORY'] = np.where(df_protect13['DMTM_TAG']=='DMTM', "DMTM", df_protect13['TERRITORY'])

df_protect13['SEGMENT'] = np.where(df_protect13['DMTM_TAG']=='DMTM', "DMTM", df_protect13['SEGMENT'])


df_protect13['BANK CATEGORY'] = np.where(df_protect13['DMTM_TAG']=='DMTM', 'DMTM', df_protect13['BANK CATEGORY'])

df_protect13['BDM'] = np.where(df_protect13['DMTM_TAG']=='DMTM', df_protect13['BDM FULL NAME'], df_protect13['BDM'])

df_protect13['BDM CODE'] = np.where(df_protect13['DMTM_TAG']=='DMTM', df_protect13['BDM CODE_db_bse_dmtm_temp'], df_protect13['BDM CODE'])

df_protect13['BAM'] = np.where(df_protect13['DMTM_TAG']=='DMTM', df_protect13['BAM FULL NAME'], df_protect13['BAM'])

df_protect13['BAM CODE'] = np.where(df_protect13['DMTM_TAG']=='DMTM', df_protect13['BAM CODE_db_bse_dmtm_temp'], df_protect13['BAM CODE'])

df_protect13['TSH'] = np.where(df_protect13['DMTM_TAG']=='DMTM', df_protect13['TSH NAME'], df_protect13['TSH'])

df_protect13['TSH CODE'] = np.where(df_protect13['DMTM_TAG']=='DMTM', df_protect13['TSH CODE_db_bse_dmtm_temp'], df_protect13['TSH CODE'])

df_protect13['TSH CODE'] = np.where(df_protect13['DMTM_TAG']=='DMTM', df_protect13['TSH CODE_db_bse_dmtm_temp'], df_protect13['TSH CODE'])

df_protect13 = df_protect13[['PolicyNumber', 'Campaign code', 'Status', 'Effective date',
       'PolicyPremium', 'Agent Code', 'Agent Name', 'Branch Code',
       'referrer code', 'ANP', 'Case Count', 'BRANCH NAME', 'DIVISION', 'AREA',
       'TERRITORY', 'BPI-AIA DIVISION', 'BPI-AIA AREA', 'BDM CODE', 'BDM',
       'BAM CODE', 'BAM', 'TSH CODE', 'TSH', 'SEGMENT', 'COUNT',
       'APPROVAL STATUS', 'FTD', 'BANK CATEGORY']]


# %% [markdown]
# ## Removing the Excepmted Policy as Requested

df_protect13 = df_protect13.loc[~df_protect13['PolicyNumber'].isin(["7013219154","7013219163"])]

df_protect13.head()

# %%
df_protect13.loc[df_protect13['BANK CATEGORY']=='DMTM'].head()

# %% [markdown]
# ## Final Transfer

# %%
PamilyaProtect_Public = df_protect13.copy()
PamilyaProtect_Public.loc[PamilyaProtect_Public['BANK CATEGORY'] == 'DMTM']

# %% [markdown]
# ## Setting the OS Directory

# %%
import os

current_dir = os.getcwd()

current_dir

current_dir = os.chdir(current_dir)

print(current_dir)

# %% [markdown]
# ## For Checking Only 

# %%
MTD_Reference_Public.head(1)

# %%
PamilyaProtect_Public.loc[PamilyaProtect_Public['PolicyNumber'] == '5101234876']

# %%
## For Date Chaning Only of MTDRef
# MTD_Reference_Public.loc[MTD_Reference_Public['EXTRACT DATE'] == '2024-07-22', 'EXTRACT DATE'] = '2024-07-19'
# ytdn_daily_1.loc[ytdn_daily_1['EXTRACT DATE'] == '2024-07-22', 'EXTRACT DATE'] = '2024-07-19'

# %% [markdown]
# # Exporting of Data

# %% [markdown]
# ## Getting the latest modification of database mastefile

# %% [markdown]
# ## To Excel Export

# %%
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import xlsxwriter

from datetime import datetime as dt

date_now = dt.now().strftime('%Y%m%d')

## Fill Na for all the Blanks
MTD_Reference_Public.fillna('')
PamilyaProtect_Public.fillna('')
ytdn_daily_1.fillna('')

Pol_tobe_Included = Monlthy_Included_Policy['CA_POLNUM'].copy()
Pol_tobe_Included = np.array(Pol_tobe_Included)
Pol_tobe_Included

MTD_Reference_Report = MTD_Reference_Public.reset_index(drop=True).copy()
MTD_Reference_Report = MTD_Reference_Report[MTD_Reference_Report['CA_POLNUM'].isin(Pol_tobe_Included)]
MTD_Reference_Report = MTD_Reference_Report.loc[MTD_Reference_Report['EXTRACT DATE'].dt.strftime('%m%Y') == dt.today().strftime('%m%Y')]

PamilyaProtect_Report = PamilyaProtect_Public.reset_index(drop=True).copy()
Daily_Reference_Report = ytdn_daily_1.reset_index(drop=True).copy()
Daily_Reference_Report = Daily_Reference_Report.loc[Daily_Reference_Report['EXTRACT DATE'].dt.strftime('%m%d%Y') == dt.today().strftime('%m%d%Y')]

YTD_Reference_Report = MTD_Reference_Public.reset_index(drop=True).copy()

Prev_Month_Policy_Report = MTD_Reference_Public.loc[MTD_Reference_Public['EXTRACT DATE'].dt.strftime('%m%Y') == (datetime.today() - relativedelta(months=1)).strftime("%m%Y")].copy()
Prev_Month_Policy_Report = Prev_Month_Policy_Report.reset_index(drop=True).copy()

# https://realpython.com/openpyxl-excel-spreadsheets-python/

# Create a Pandas ExcelWriter object
with pd.ExcelWriter(f"MTD_{date_now} Reference with T1.xlsx", engine='xlsxwriter', datetime_format="mm/dd/yyyy", date_format="mm/dd/yyyy") as writer:
# Write the DataFrames to Excel
    MTD_Reference_Report.to_excel(writer, sheet_name='MTD Reference', index=False)
    YTD_Reference_Report.to_excel(writer, sheet_name='YTD Reference', index=False)
    Prev_Month_Policy_Report.to_excel(writer, sheet_name='Previous Month Policy Report', index=False)
    PamilyaProtect_Report.to_excel(writer, sheet_name='PP FTM', index=False)
    Daily_Reference_Report.to_excel(writer, sheet_name='FTD Reference', index=False)

    workbook = writer.book

    #workbook.formats[0].set_font_size(10)
    #font_size_custom = workbook.add_format({'font_size':8})
    
    #border_fmt = workbook.add_format({'font_size':8, 'font_name':'Arial,','bottom':1, 'top':1, 'left':1, 'right':1})
    border_fmt = workbook.add_format({'border':1})

    sheet1 = writer.sheets['MTD Reference']
    sheet1.set_tab_color('red')
    (max_row, max_col) = MTD_Reference_Report.shape
    sheet1.set_column(0, max_col, 30)

    sheet2 = writer.sheets['YTD Reference']
    sheet2.set_tab_color('red')
    (max_row, max_col) = YTD_Reference_Report.shape
    sheet2.set_column(0, max_col, 30)

    sheet2_1 = writer.sheets['Previous Month Policy Report']
    sheet2_1.set_tab_color('red')
    (max_row, max_col) = Prev_Month_Policy_Report.shape
    sheet2_1.set_column(0, max_col, 30)

    sheet3 = writer.sheets['PP FTM']
    sheet3.set_tab_color('red')
    (max_row, max_col) = PamilyaProtect_Report.shape
    sheet3.set_column(0, max_col, 30)

    sheet4 = writer.sheets['FTD Reference']
    sheet4.set_tab_color('red')
    (max_row, max_col) = Daily_Reference_Report.shape
    sheet4.set_column(0, max_col, 30)



    # Add a header format.
    header_format = workbook.add_format({'bold': True,
                                        'bottom': 2,
                                        'font_color':'white',
                                        'bg_color': '#0504aa'})

    # Write the column headers with the defined format.
    sheet_list = [sheet1, sheet2, sheet2_1, sheet3, sheet4]
    sheet_df = [MTD_Reference_Report, YTD_Reference_Report, Prev_Month_Policy_Report, PamilyaProtect_Report, Daily_Reference_Report]

    for sheet_val in range(len(sheet_list)):
        print(sheet_val, " loop")
        for col_num, value in enumerate(sheet_df[sheet_val].columns.values):
            sheet_list[sheet_val].write(0, col_num, value, header_format)
            sheet_list[sheet_val].conditional_format(xlsxwriter.utility.xl_range(0, 0, len(sheet_df[sheet_val]), len(sheet_df[sheet_val].columns)-1), {'type': 'no_errors', 'format': border_fmt})
    
print(os.getcwd())

# %% [markdown]
# # Auto Send Email

# %%


# %%
import os
import win32com.client as win32
import glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime as dt

query = '''  SELECT TOP(1) OBJECT_NAME(OBJECT_ID) AS DatabaseName, CAST(last_user_update AS NVARCHAR) AS [last_user_update]
FROM sys.dm_db_index_usage_stats WHERE last_user_update IS NOT NULL
ORDER BY last_user_update DESC  '''

db_update = sql_connection(query, 'db_MasterFile_2024')

path_file_upload = os.getcwd()

Date = dt.now().strftime('%B %d, %Y')

olApp = win32.Dispatch('Outlook.Application')
olNS = olApp.GetNamespace("MAPI")

mailItem = olApp.CreateItem(0)
mailItem.Subject = f'MTD Reference - {Date}'
mailItem.BodyFormat = 1
mailItem.HTMLBody = f'''
<!DOCTYPE html>
    <html>
    <body>
    <div><p>Dear All,</p></div>
    <div><p>Please see MTD Reference File.</p></div>
    <div>
    <p> Updated Database as of {db_update.iat[0,1]} </p> 
        <p> Have added an OLD/NEW tenure  for BSE  beside BSE code Column.
            Pamilya Protect is as of {Date}.
        </p>
    </div>

<div><p>Best Regards,</p></div>
</body>
</html>
''' 
#mailItem.To = 'ashnergerald.novilla@aia.com'
mailItem.To = 'joshuaaudie-ja.depositario@aia.com'
mailItem.cc = 'jonathan-j.tomas@aia.com; christian.simpao@aia.com'
mailItem.Attachments.Add(path_file_upload+f'\\MTD_{date_now} Reference with T1.xlsx')
#mailItem.Attachments.Add(ytdn)

mailItem.SentOnBehalfOfName = "philippines.bplac.bas.incentive-prod@aia.com"
#mailItem.Save()
mailItem.Send()
# mailItem.Display()

# %% [markdown]
# # Upload to SQL

# %%
import pandas as pd
from sqlalchemy import create_engine
from datetime import datetime
pd.set_option('display.max_colwidth', None)

server = 'PPBWDLC0SG7A1'
database = 'db_MTD_Ref_2024'
username = 'admin'
password = 'Openlab@123'

MTD_Reference_Report = MTD_Reference_Public.reset_index(drop=True).copy()
PamilyaProtect_Report = PamilyaProtect_Public.reset_index(drop=True).copy()
Daily_Reference_Report = ytdn_daily_1.reset_index(drop=True).copy()

connection_string = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
engine = create_engine('mssql+pyodbc:///?odbc_connect={}'.format(connection_string))

MTD_Reference_Report['DateUpload'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
PamilyaProtect_Report['DateUpload'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
Daily_Reference_Report['DateUpload'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# %%
now_year = datetime.now().strftime('%Y')
try:    
    db_mtdref_old.to_sql(f'tbl_Prev_MTD_Reference_Report_{now_year}', engine, if_exists='replace', index=False)
    print("Success Upload")
except:
    raise ValueError('SQL File Not Upload.')

try:
    MTD_Reference_Report.to_sql(f'tbl_MTD_Reference_Report_{now_year}', engine, if_exists='replace', index=False)
    PamilyaProtect_Report.to_sql(f'tbl_PamilyaProtect_Report_{now_year}', engine, if_exists='replace', index=False)
    Daily_Reference_Report.to_sql(f'tbl_Daily_Reference_Report_{now_year}', engine, if_exists='replace', index=False)
    print("Success Upload")
except:
    raise ValueError('SQL File Not Upload.')

# try:
#     PamilyaProtect_Report.to_sql('tbl_PamilyaProtect_Report_2024', engine, if_exists='replace', index=False)
#     print("Success PamilyaProtect_Report")
# except:
#     raise ValueError('SQL File Not Upload.')

# try:
#     Daily_Reference_Report.to_sql('tbl_Daily_Reference_Report_2024', engine, if_exists='replace', index=False)
#     print("Success Daily_Reference_Report")

# except:
#     raise ValueError('SQL File Not Upload.')

del(MTD_Reference_Report, PamilyaProtect_Report, Daily_Reference_Report) 

# %%
try:
    subprocess.call("TASKKILL /F /IM acslaunch_win-32.exe", shell=True)
    dest = shutil.move(source, destination) 
    sleep(5)
except:
    pass
    
try:
    subprocess.call("TASKKILL /F /IM acslaunch_win-64.exe", shell=True)
    dest = shutil.move(source, destination) 
    sleep(5)
except:
    pass

# %% [markdown]
# # Upload to SQL Jonathan Customs

# %%
# import os
# # os.chdir(r'\\palp3r7cfis08\BPLAC MIS LIBRARY\Git_Codes\Temp_tables')
# os.chdir(r'W:\Git_Codes\Temp_tables')
# import YTD_Subs_update_2024
import subprocess

filepath=r'\\palp3r7cfis08\BPLAC MIS LIBRARY\Git_Codes\Temp_tables\YTD_Subs_Update.bat'
p = subprocess.Popen(filepath, shell=True, stdout = subprocess.PIPE)

stdout, stderr = p.communicate()
print (p.returncode) # is 0 if success

